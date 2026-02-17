import os
import json
import re
from PyQt5 import QtWidgets, QtCore, QtGui
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell


# =============================================================================
# COLUMN DEFINITIONS - Separate basic (Techniker) from technical (Entwickler)
# =============================================================================

# Columns that are specific to certain classes only
CLASS_SPECIFIC_COLUMNS = {
    'fahrtrichtung': ['signal'],  # Only for signal class
}

# Techniker: Only the 3 columns shown in the main UI table
TECHNIKER_COLUMNS = {
    'anchor_text': 'Text/Nummer',
    'coord_text': 'Koordinatentext',
    'fahrtrichtung': 'Fahrtrichtung',  # Only shown for signal class
}

# Basic columns (more than Techniker but still user-friendly)
BASIC_COLUMNS = {
    'anchor_text': 'Text/Nummer',
    'coord_text': 'Koordinatentext',
    'coord_value': 'Koordinatenwert',
    'gi_gl': 'GI/GL',
    'page': 'Seite',
    'fahrtrichtung': 'Fahrtrichtung',
    'weichen_coordinates': 'Weichen Koordinaten',
    'notes': 'Notizen',
}

TECHNICAL_COLUMNS = {
    # Columns for developers/debugging
    'conf': 'Konfidenz',
    'color': 'Farbe',
    'ax1': 'Anker X1', 'ay1': 'Anker Y1',
    'ax2': 'Anker X2', 'ay2': 'Anker Y2',
    'cx1': 'Koord X1', 'cy1': 'Koord Y1',
    'cx2': 'Koord X2', 'cy2': 'Koord Y2',
    'angle': 'Winkel (normalisiert)',
    'angle_raw': 'Winkel (roh)',
    'obb_cx': 'OBB Center X',
    'obb_cy': 'OBB Center Y',
    'obb_w': 'OBB Breite',
    'obb_h': 'OBB Höhe',
}


class SimpleExcelExportDialog(QtWidgets.QDialog):
    """
    Simplified Excel export dialog with:
    - Tree view for class/column selection
    - Live data preview
    - Statistics panel
    - Export summary
    """

    def __init__(self, parent: 'WorkspaceWidget'):
        super().__init__(parent)
        self.workspace = parent
        self.setWindowTitle("Excel Export")
        self.resize(1000, 550)
        self._cached_scan_result = None  # Cache for target area scan

        self.setWindowFlags(
            QtCore.Qt.Window |
            QtCore.Qt.WindowMinimizeButtonHint |
            QtCore.Qt.WindowMaximizeButtonHint |
            QtCore.Qt.WindowCloseButtonHint
        )

        self._build_ui()
        self._populate_tree()
        QtCore.QTimer.singleShot(100, self._update_preview)

    def _build_ui(self):
        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setSpacing(6)

        # ============================================================
        # TOOLBAR
        # ============================================================
        toolbar_frame = QtWidgets.QFrame()
        toolbar_frame.setFrameStyle(QtWidgets.QFrame.StyledPanel)
        toolbar_layout = QtWidgets.QHBoxLayout(toolbar_frame)
        toolbar_layout.setContentsMargins(8, 4, 8, 4)
        toolbar_layout.setSpacing(4)

        # Presets
        toolbar_layout.addWidget(QtWidgets.QLabel("<b>Vorlagen:</b>"))
        self.btn_techniker = QtWidgets.QPushButton(" Techniker")
        self.btn_techniker.setToolTip("Nur die 3 Basis-Spalten")
        self.btn_techniker.clicked.connect(self._select_techniker_preset)
        toolbar_layout.addWidget(self.btn_techniker)

        self.btn_entwickler = QtWidgets.QPushButton(" Entwickler")
        self.btn_entwickler.setToolTip("Alle Spalten inkl. technische")
        self.btn_entwickler.clicked.connect(self._select_entwickler_preset)
        toolbar_layout.addWidget(self.btn_entwickler)

        toolbar_layout.addWidget(self._create_separator())

        # Quick selection
        toolbar_layout.addWidget(QtWidgets.QLabel("<b>Auswahl:</b>"))
        btn_select_all = QtWidgets.QPushButton(" Alle")
        btn_select_all.setToolTip("Alle Klassen und Spalten auswählen")
        btn_select_all.clicked.connect(self._select_all)
        toolbar_layout.addWidget(btn_select_all)

        btn_select_none = QtWidgets.QPushButton(" Keine")
        btn_select_none.setToolTip("Alle abwählen")
        btn_select_none.clicked.connect(self._select_none)
        toolbar_layout.addWidget(btn_select_none)

        toolbar_layout.addWidget(self._create_separator())

        # Help
        toolbar_layout.addStretch()
        help_btn = QtWidgets.QPushButton(" Hilfe")
        help_btn.clicked.connect(self._show_help)
        toolbar_layout.addWidget(help_btn)

        main_layout.addWidget(toolbar_frame)

        # ============================================================
        # MAIN CONTENT: Tree + Preview side by side
        # ============================================================
        content_splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        content_splitter.setHandleWidth(6)
        content_splitter.setChildrenCollapsible(False)

        # LEFT: Tree view for class/column selection
        left_widget = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)

        # Tree widget
        self.tree = QtWidgets.QTreeWidget()
        self.tree.setHeaderLabels(["Auswahl", "Anzahl"])
        self.tree.setColumnWidth(0, 250)
        self.tree.setColumnWidth(1, 60)
        self.tree.itemChanged.connect(self._on_tree_item_changed)
        left_layout.addWidget(self.tree)

        # Expand/collapse buttons
        tree_btn_layout = QtWidgets.QHBoxLayout()
        btn_expand = QtWidgets.QPushButton("▼ Aufklappen")
        btn_expand.clicked.connect(self.tree.expandAll)
        tree_btn_layout.addWidget(btn_expand)

        btn_collapse = QtWidgets.QPushButton("▶ Zuklappen")
        btn_collapse.clicked.connect(self.tree.collapseAll)
        tree_btn_layout.addWidget(btn_collapse)

        tree_btn_layout.addStretch()
        left_layout.addLayout(tree_btn_layout)

        content_splitter.addWidget(left_widget)

        # RIGHT: Data preview
        right_widget = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # Preview header with controls
        preview_header = QtWidgets.QHBoxLayout()
        self.preview_label = QtWidgets.QLabel("<b>Vorschau:</b> <small>(Spalten ziehen zum Umsortieren, Rechtsklick für Optionen)</small>")
        preview_header.addWidget(self.preview_label)
        preview_header.addStretch()

        self.btn_exclude_rows = QtWidgets.QPushButton("Zeilen ausschließen")
        self.btn_exclude_rows.setEnabled(False)
        self.btn_exclude_rows.setToolTip("Markierte Zeilen vom Export ausschließen")
        self.btn_exclude_rows.clicked.connect(self._exclude_selected_rows)
        preview_header.addWidget(self.btn_exclude_rows)

        self.btn_include_all = QtWidgets.QPushButton("Zurücksetzen")
        self.btn_include_all.setToolTip("Alle Ausschlüsse zurücksetzen")
        self.btn_include_all.clicked.connect(self._include_all)
        preview_header.addWidget(self.btn_include_all)

        right_layout.addLayout(preview_header)

        # Preview table with drag & drop column reordering
        self.preview_table = QtWidgets.QTableWidget()
        self.preview_table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.preview_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.preview_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        self.preview_table.horizontalHeader().setSectionsMovable(True)  # Enable column drag
        self.preview_table.horizontalHeader().sectionClicked.connect(self._on_column_header_clicked)
        self.preview_table.horizontalHeader().setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.preview_table.horizontalHeader().customContextMenuRequested.connect(self._show_column_context_menu)
        self.preview_table.itemSelectionChanged.connect(self._on_preview_selection_changed)
        right_layout.addWidget(self.preview_table)

        # Track excluded columns and column renames
        self.excluded_columns = set()
        self.column_renames = {}  # {original_name: new_name}

        # Row count and excluded info
        row_info_layout = QtWidgets.QHBoxLayout()
        self.row_count_label = QtWidgets.QLabel("")
        row_info_layout.addWidget(self.row_count_label)
        row_info_layout.addStretch()
        self.excluded_label = QtWidgets.QLabel("")
        row_info_layout.addWidget(self.excluded_label)
        right_layout.addLayout(row_info_layout)

        # Track excluded rows
        self.excluded_row_indices = set()

        content_splitter.addWidget(right_widget)
        content_splitter.setSizes([300, 700])

        # ============================================================
        # MAIN VERTICAL SPLITTER (content on top, options on bottom)
        # ============================================================
        main_splitter = QtWidgets.QSplitter(QtCore.Qt.Vertical)
        main_splitter.setHandleWidth(6)
        main_splitter.setChildrenCollapsible(False)
        main_splitter.addWidget(content_splitter)

        # Bottom panel container
        bottom_widget = QtWidgets.QWidget()
        bottom_layout = QtWidgets.QVBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        bottom_layout.setSpacing(4)

        # ============================================================
        # STATISTICS & SUMMARY PANEL
        # ============================================================
        stats_frame = QtWidgets.QFrame()
        stats_frame.setFrameStyle(QtWidgets.QFrame.StyledPanel)
        stats_layout = QtWidgets.QHBoxLayout(stats_frame)
        stats_layout.setContentsMargins(10, 6, 10, 6)

        # Left: Class statistics
        stats_left = QtWidgets.QVBoxLayout()
        stats_left.addWidget(QtWidgets.QLabel("<b>Statistik:</b>"))
        self.stats_label = QtWidgets.QLabel("Keine Daten ausgewählt")
        self.stats_label.setWordWrap(True)
        stats_left.addWidget(self.stats_label)
        stats_layout.addLayout(stats_left, 1)

        # Separator
        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.VLine)
        stats_layout.addWidget(separator)

        # Right: Export summary
        stats_right = QtWidgets.QVBoxLayout()
        stats_right.addWidget(QtWidgets.QLabel("<b>Export-Zusammenfassung:</b>"))
        self.summary_label = QtWidgets.QLabel("Wählen Sie Daten zum Exportieren")
        self.summary_label.setWordWrap(True)
        stats_right.addWidget(self.summary_label)
        stats_layout.addLayout(stats_right, 1)

        # Separator
        separator2 = QtWidgets.QFrame()
        separator2.setFrameShape(QtWidgets.QFrame.VLine)
        stats_layout.addWidget(separator2)

        # Format selection
        format_layout = QtWidgets.QVBoxLayout()
        format_layout.addWidget(QtWidgets.QLabel("<b>Format:</b>"))
        self.format_combo = QtWidgets.QComboBox()
        self.format_combo.addItems(["Excel (.xlsx)", "CSV (.csv)", "JSON (.json)"])
        format_layout.addWidget(self.format_combo)
        stats_layout.addLayout(format_layout)

        bottom_layout.addWidget(stats_frame)

        # ============================================================
        # EXPORT TARGET
        # ============================================================
        target_frame = QtWidgets.QFrame()
        target_frame.setFrameStyle(QtWidgets.QFrame.StyledPanel | QtWidgets.QFrame.Raised)
        target_main_layout = QtWidgets.QVBoxLayout(target_frame)
        target_main_layout.setContentsMargins(10, 10, 10, 10)
        target_main_layout.setSpacing(8)

        # Top row: Target selection
        target_row1 = QtWidgets.QHBoxLayout()
        target_row1.addWidget(QtWidgets.QLabel("<b>Ziel:</b>"))

        self.radio_new = QtWidgets.QRadioButton("Neue Datei")
        self.radio_new.setChecked(True)
        self.radio_new.toggled.connect(self._on_target_changed)
        target_row1.addWidget(self.radio_new)

        self.radio_existing = QtWidgets.QRadioButton("Bestehende:")
        self.radio_existing.toggled.connect(self._on_target_changed)
        target_row1.addWidget(self.radio_existing)

        self.existing_file_edit = QtWidgets.QLineEdit()
        self.existing_file_edit.setPlaceholderText("Datei wählen...")
        self.existing_file_edit.setReadOnly(True)
        self.existing_file_edit.setMaximumWidth(180)
        self.existing_file_edit.setEnabled(False)
        target_row1.addWidget(self.existing_file_edit)

        self.btn_browse = QtWidgets.QPushButton("...")
        self.btn_browse.setMaximumWidth(30)
        self.btn_browse.setEnabled(False)
        self.btn_browse.clicked.connect(self._browse_file)
        target_row1.addWidget(self.btn_browse)

        target_row1.addStretch()
        target_main_layout.addLayout(target_row1)

        # Second row: Position settings (only for existing file)
        self.position_widget = QtWidgets.QWidget()
        target_row2 = QtWidgets.QHBoxLayout(self.position_widget)
        target_row2.setContentsMargins(0, 0, 0, 0)

        target_row2.addWidget(QtWidgets.QLabel("Blatt:"))
        self.sheet_combo = QtWidgets.QComboBox()
        self.sheet_combo.setMinimumWidth(100)
        target_row2.addWidget(self.sheet_combo)

        target_row2.addWidget(QtWidgets.QLabel("Start-Spalte:"))
        self.pos_column = QtWidgets.QLineEdit("A")
        self.pos_column.setMaximumWidth(40)
        target_row2.addWidget(self.pos_column)

        target_row2.addWidget(QtWidgets.QLabel("Start-Zeile:"))
        self.pos_row = QtWidgets.QSpinBox()
        self.pos_row.setMinimum(1)
        self.pos_row.setMaximum(1048576)
        self.pos_row.setValue(1)
        self.pos_row.setMaximumWidth(70)
        target_row2.addWidget(self.pos_row)

        self.check_append = QtWidgets.QCheckBox("Auto-Anfügen (nach letzter Zeile)")
        target_row2.addWidget(self.check_append)

        # Mini Excel preview (before stretch so it's more to the left)
        self.excel_preview = QtWidgets.QLabel()
        self.excel_preview.setStyleSheet("font-family: monospace; font-size: 12px; padding: 5px; border: 1px solid gray; border-radius: 3px;")
        self.excel_preview.setMinimumWidth(150)
        target_row2.addWidget(self.excel_preview)

        target_row2.addStretch()

        self.position_widget.setVisible(False)
        target_main_layout.addWidget(self.position_widget)

        # Position explanation label
        self.position_hint = QtWidgets.QLabel("")
        target_main_layout.addWidget(self.position_hint)

        # Connect signals to update hint
        self.pos_column.textChanged.connect(self._update_position_hint)
        self.pos_row.valueChanged.connect(self._update_position_hint)

        bottom_layout.addWidget(target_frame)

        # ============================================================
        # OPTIONS - Two rows for better organization
        # ============================================================
        options_frame = QtWidgets.QFrame()
        options_frame.setFrameStyle(QtWidgets.QFrame.StyledPanel)
        options_main = QtWidgets.QVBoxLayout(options_frame)
        options_main.setContentsMargins(10, 6, 10, 6)
        options_main.setSpacing(6)

        # Row 1: Basic options
        options_row1 = QtWidgets.QHBoxLayout()
        options_row1.addWidget(QtWidgets.QLabel("<b>Optionen:</b>"))

        self.check_include_header = QtWidgets.QCheckBox("Mit Überschriften")
        self.check_include_header.setChecked(True)
        self.check_include_header.setToolTip("Spaltennamen in erster Zeile")
        options_row1.addWidget(self.check_include_header)

        self.check_auto_width = QtWidgets.QCheckBox("Spaltenbreite anpassen")
        self.check_auto_width.setChecked(True)
        options_row1.addWidget(self.check_auto_width)

        self.check_freeze_header = QtWidgets.QCheckBox("Kopfzeile fixieren")
        self.check_freeze_header.setChecked(True)
        self.check_freeze_header.setToolTip("Kopfzeile beim Scrollen sichtbar")
        options_row1.addWidget(self.check_freeze_header)

        self.check_apply_filters = QtWidgets.QCheckBox("Filter anwenden")
        self.check_apply_filters.setChecked(True)
        options_row1.addWidget(self.check_apply_filters)

        options_row1.addWidget(self._create_separator())

        options_row1.addWidget(QtWidgets.QLabel("Sortierung:"))
        self.sort_combo = QtWidgets.QComboBox()
        self.sort_combo.addItems([
            "Keine",
            "Klasse (A→Z)", "Klasse (Z→A)",
            "Seite (1→9)", "Seite (9→1)",
            "Koordinate (A→Z)", "Koordinate (Z→A)"
        ])
        options_row1.addWidget(self.sort_combo)

        options_row1.addStretch()
        options_main.addLayout(options_row1)

        # Row 2: Format-specific options
        options_row2 = QtWidgets.QHBoxLayout()

        # Grouping
        options_row2.addWidget(QtWidgets.QLabel("Gruppierung:"))
        self.grouping_combo = QtWidgets.QComboBox()
        self.grouping_combo.addItems([
            "Keine",
            "Nach Klasse (Abschnitte)",
            "Nach Klasse (separate Blätter)"
        ])
        self.grouping_combo.setToolTip("Wie Daten gruppiert werden")
        options_row2.addWidget(self.grouping_combo)

        options_row2.addWidget(self._create_separator())

        # Empty cells
        options_row2.addWidget(QtWidgets.QLabel("Leere Zellen:"))
        self.empty_cell_combo = QtWidgets.QComboBox()
        self.empty_cell_combo.addItems(["Leer lassen", "Mit '-' füllen", "Mit 'N/A' füllen"])
        options_row2.addWidget(self.empty_cell_combo)

        options_row2.addWidget(self._create_separator())

        # Number format
        options_row2.addWidget(QtWidgets.QLabel("Zahlen:"))
        self.number_format_combo = QtWidgets.QComboBox()
        self.number_format_combo.addItems(["Original", "2 Dezimalen", "Ganzzahl"])
        options_row2.addWidget(self.number_format_combo)

        # Coordinate scaling
        options_row2.addWidget(QtWidgets.QLabel("Koordinaten:"))
        self.coord_scale_combo = QtWidgets.QComboBox()
        self.coord_scale_combo.addItems([
            "Original",
            "×1000 (m→mm)",
            "×100 (m→cm)",
            "×10",
            "÷1000 (mm→m)",
            "÷100 (cm→m)"
        ])
        self.coord_scale_combo.setToolTip("Skalierung für Koordinaten-Spalten (x, y, width, height)")
        options_row2.addWidget(self.coord_scale_combo)

        options_row2.addWidget(self._create_separator())

        # CSV options (only shown when CSV selected)
        self.csv_options_widget = QtWidgets.QWidget()
        csv_layout = QtWidgets.QHBoxLayout(self.csv_options_widget)
        csv_layout.setContentsMargins(0, 0, 0, 0)
        csv_layout.addWidget(QtWidgets.QLabel("CSV-Trennzeichen:"))
        self.csv_separator_combo = QtWidgets.QComboBox()
        self.csv_separator_combo.addItems(["Semikolon (;)", "Komma (,)", "Tab", "Pipe (|)"])
        self.csv_separator_combo.setToolTip("Semikolon für deutsches Excel empfohlen")
        csv_layout.addWidget(self.csv_separator_combo)
        options_row2.addWidget(self.csv_options_widget)

        # JSON options (only shown when JSON selected)
        self.json_options_widget = QtWidgets.QWidget()
        json_layout = QtWidgets.QHBoxLayout(self.json_options_widget)
        json_layout.setContentsMargins(0, 0, 0, 0)
        json_layout.addWidget(QtWidgets.QLabel("JSON-Struktur:"))
        self.json_orient_combo = QtWidgets.QComboBox()
        self.json_orient_combo.addItems([
            "Records (Liste von Objekten)",
            "Nach Klasse gruppiert",
            "Tabelle (columns/data)",
            "Index (dict of dicts)"
        ])
        self.json_orient_combo.setToolTip("Struktur der JSON-Ausgabe")
        json_layout.addWidget(self.json_orient_combo)
        self.json_options_widget.setVisible(False)
        options_row2.addWidget(self.json_options_widget)

        options_row2.addStretch()
        options_main.addLayout(options_row2)

        bottom_layout.addWidget(options_frame)

        # Connect format change to show/hide CSV options
        self.format_combo.currentTextChanged.connect(self._on_format_changed)

        # Connect format combo to update summary
        self.format_combo.currentTextChanged.connect(self._update_preview)

        # ============================================================
        # ACTION BUTTONS
        # ============================================================
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addStretch()

        self.btn_export = QtWidgets.QPushButton("Exportieren")
        self.btn_export.clicked.connect(self._on_export)
        button_layout.addWidget(self.btn_export)

        self.btn_cancel = QtWidgets.QPushButton("Abbrechen")
        self.btn_cancel.clicked.connect(self.reject)
        button_layout.addWidget(self.btn_cancel)

        bottom_layout.addLayout(button_layout)

        # Add bottom panel to main splitter
        main_splitter.addWidget(bottom_widget)
        main_splitter.setSizes([400, 150])

        main_layout.addWidget(main_splitter)

    def _create_separator(self):
        """Create a vertical line separator for toolbars"""
        sep = QtWidgets.QFrame()
        sep.setFrameShape(QtWidgets.QFrame.VLine)
        sep.setFrameShadow(QtWidgets.QFrame.Sunken)
        return sep

    def _select_all(self):
        """Select all classes and columns"""
        self.tree.blockSignals(True)
        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            class_item.setCheckState(0, QtCore.Qt.Checked)
            class_item.setHidden(False)
            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                col_item.setCheckState(0, QtCore.Qt.Checked)
                col_item.setHidden(False)
        self.tree.blockSignals(False)
        self._update_preview()

    def _select_none(self):
        """Deselect all classes and columns"""
        self.tree.blockSignals(True)
        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            class_item.setCheckState(0, QtCore.Qt.Unchecked)
            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                col_item.setCheckState(0, QtCore.Qt.Unchecked)
        self.tree.blockSignals(False)
        self._update_preview()

    def _on_format_changed(self, format_text):
        """Show/hide format-specific options"""
        is_csv = 'CSV' in format_text
        is_json = 'JSON' in format_text
        is_excel = not is_csv and not is_json

        # Show/hide format-specific options
        self.csv_options_widget.setVisible(is_csv)
        self.json_options_widget.setVisible(is_json)

        # Disable Excel-only options for CSV/JSON
        self.check_freeze_header.setEnabled(is_excel)
        self.check_auto_width.setEnabled(is_excel)

        # Reset grouping if separate sheets selected for non-Excel
        if 'separate Blätter' in self.grouping_combo.currentText() and not is_excel:
            self.grouping_combo.setCurrentIndex(0)  # Reset to "Keine"

        # Disable "existing file" option for JSON (only new file supported)
        if is_json:
            self.radio_new.setChecked(True)
            self.radio_existing.setEnabled(False)
        else:
            self.radio_existing.setEnabled(True)

    def _populate_tree(self):
        """Populate tree with classes and columns"""
        self.tree.clear()

        if self.workspace.df_all is None or self.workspace.df_all.empty:
            return

        df = self.workspace.df_all
        if 'cls' not in df.columns:
            return
        classes = sorted(df['cls'].unique())

        # Move coordinate to end
        if 'coordinate' in classes:
            classes.remove('coordinate')
            classes.append('coordinate')

        for cls_name in classes:
            df_class = df[df['cls'] == cls_name]
            count = len(df_class)

            # Class item - start unchecked (user selects what they want)
            class_item = QtWidgets.QTreeWidgetItem()
            class_item.setText(0, f"{cls_name}")
            class_item.setText(1, f"({count})")
            class_item.setFlags(class_item.flags() | QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsAutoTristate)
            class_item.setCheckState(0, QtCore.Qt.Unchecked)
            class_item.setData(0, QtCore.Qt.UserRole, ('class', cls_name))

            # Get available columns for this class
            available_cols = self._get_class_columns(cls_name, df_class)

            # Add column items - all unchecked by default
            for col_key, col_display in available_cols.items():
                is_technical = col_key in TECHNICAL_COLUMNS and col_key not in BASIC_COLUMNS

                col_item = QtWidgets.QTreeWidgetItem()
                col_item.setText(0, col_display)
                col_item.setFlags(col_item.flags() | QtCore.Qt.ItemIsUserCheckable)

                # All unchecked by default - user picks what to export
                col_item.setCheckState(0, QtCore.Qt.Unchecked)
                if is_technical:
                    col_item.setForeground(0, QtGui.QColor(128, 128, 128))  # Gray for technical

                col_item.setData(0, QtCore.Qt.UserRole, ('column', col_key, cls_name))
                class_item.addChild(col_item)

            self.tree.addTopLevelItem(class_item)

        # Start collapsed - user expands what they need
        self.tree.collapseAll()

    def _get_class_columns(self, cls_name: str, df_class) -> dict:
        """Get available columns for a class"""
        all_cols = {**BASIC_COLUMNS, **TECHNICAL_COLUMNS}

        # Filter to existing columns and class-specific columns
        result = {}
        for col_key, col_display in all_cols.items():
            if col_key not in df_class.columns:
                continue

            # Check if this is a class-specific column
            if col_key in CLASS_SPECIFIC_COLUMNS:
                allowed_classes = CLASS_SPECIFIC_COLUMNS[col_key]
                if cls_name not in allowed_classes:
                    continue  # Skip this column for this class

            result[col_key] = col_display

        return result

    def _on_tree_item_changed(self, item, column):
        """Handle tree item check state change"""
        # Block signals to prevent recursion
        self.tree.blockSignals(True)

        data = item.data(0, QtCore.Qt.UserRole)
        if data and data[0] == 'class':
            # Class item changed
            check_state = item.checkState(0)

            if check_state == QtCore.Qt.Unchecked:
                # Save current column selection before unchecking
                saved_cols = []
                for i in range(item.childCount()):
                    child = item.child(i)
                    if child.checkState(0) == QtCore.Qt.Checked:
                        col_data = child.data(0, QtCore.Qt.UserRole)
                        if col_data:
                            saved_cols.append(col_data[1])
                # Store in item data
                item.setData(0, QtCore.Qt.UserRole + 1, saved_cols)
                # Uncheck all children
                for i in range(item.childCount()):
                    item.child(i).setCheckState(0, QtCore.Qt.Unchecked)

            elif check_state == QtCore.Qt.Checked:
                # Restore previous selection if available
                saved_cols = item.data(0, QtCore.Qt.UserRole + 1)
                if saved_cols:
                    # Restore saved selection
                    for i in range(item.childCount()):
                        child = item.child(i)
                        col_data = child.data(0, QtCore.Qt.UserRole)
                        if col_data and col_data[1] in saved_cols:
                            child.setCheckState(0, QtCore.Qt.Checked)
                        else:
                            child.setCheckState(0, QtCore.Qt.Unchecked)
                else:
                    # First time - select default (Techniker) columns
                    for i in range(item.childCount()):
                        child = item.child(i)
                        col_data = child.data(0, QtCore.Qt.UserRole)
                        if col_data and col_data[1] in TECHNIKER_COLUMNS:
                            child.setCheckState(0, QtCore.Qt.Checked)
                        else:
                            child.setCheckState(0, QtCore.Qt.Unchecked)

        self.tree.blockSignals(False)

        # Update preview with delay
        if not hasattr(self, '_preview_timer'):
            self._preview_timer = QtCore.QTimer()
            self._preview_timer.setSingleShot(True)
            self._preview_timer.timeout.connect(self._update_preview)
        self._preview_timer.start(200)

    def _select_techniker_preset(self):
        """Techniker mode: Show ONLY the 3 main UI columns, hide all others"""
        self.tree.blockSignals(True)

        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            class_item.setCheckState(0, QtCore.Qt.Checked)
            class_item.setHidden(False)

            has_visible_cols = False
            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                col_data = col_item.data(0, QtCore.Qt.UserRole)
                if col_data and col_data[0] == 'column':
                    col_key = col_data[1]
                    # Only show the 3 Techniker columns
                    if col_key in TECHNIKER_COLUMNS:
                        col_item.setHidden(False)
                        col_item.setCheckState(0, QtCore.Qt.Checked)
                        has_visible_cols = True
                    else:
                        col_item.setHidden(True)
                        col_item.setCheckState(0, QtCore.Qt.Unchecked)

            # Hide class if it has no visible columns
            if not has_visible_cols:
                class_item.setHidden(True)

        self.tree.blockSignals(False)
        self._update_preview()

    def _select_entwickler_preset(self):
        """Entwickler mode: Show ALL columns"""
        self.tree.blockSignals(True)

        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            class_item.setHidden(False)
            class_item.setCheckState(0, QtCore.Qt.Checked)

            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                col_item.setHidden(False)
                col_item.setCheckState(0, QtCore.Qt.Checked)

        self.tree.blockSignals(False)
        self._update_preview()

    def _on_preview_selection_changed(self):
        """Handle preview table selection change"""
        selected_rows = self.preview_table.selectionModel().selectedRows()
        self.btn_exclude_rows.setEnabled(len(selected_rows) > 0)

    def _on_column_header_clicked(self, logical_index):
        """Toggle column exclusion when header is clicked"""
        # Skip the first column (Klasse reference column)
        if logical_index == 0:
            return

        # Get the column name from the header
        header_item = self.preview_table.horizontalHeaderItem(logical_index)
        if not header_item:
            return

        col_name = header_item.text()

        # Toggle exclusion
        if col_name in self.excluded_columns:
            self.excluded_columns.discard(col_name)
        else:
            self.excluded_columns.add(col_name)

        # Update visual display
        self._update_excluded_column_display()
        self._update_export_summary()

    def _show_column_context_menu(self, pos):
        """Show context menu for column operations"""
        header = self.preview_table.horizontalHeader()
        logical_index = header.logicalIndexAt(pos)

        if logical_index < 0:
            return

        # Get clicked column name
        header_item = self.preview_table.horizontalHeaderItem(logical_index)
        if not header_item:
            return

        col_name = header_item.text()
        is_reference_col = logical_index == 0  # Klasse reference column

        menu = QtWidgets.QMenu(self)

        if not is_reference_col:
            # Toggle exclude/include
            if col_name in self.excluded_columns:
                action_toggle = menu.addAction(f" '{col_name}' einschließen")
            else:
                action_toggle = menu.addAction(f" '{col_name}' ausschließen")
            action_toggle.triggered.connect(lambda: self._on_column_header_clicked(logical_index))

            menu.addSeparator()

            # Rename column for export
            action_rename = menu.addAction(f" '{col_name}' umbenennen...")
            action_rename.triggered.connect(lambda: self._rename_column(col_name, logical_index))

            # Remove column from tree (uncheck in tree)
            action_remove = menu.addAction(f" '{col_name}' aus Export entfernen")
            action_remove.triggered.connect(lambda: self._remove_column_from_tree(col_name))

        menu.addSeparator()

        # Add columns submenu - show available columns that are not currently shown
        add_menu = menu.addMenu(" Spalte hinzufügen")
        available_cols = self._get_addable_columns()
        if available_cols:
            for col_key, col_display in available_cols.items():
                action = add_menu.addAction(col_display)
                action.triggered.connect(lambda checked, k=col_key: self._add_column_from_menu(k))
        else:
            no_cols = add_menu.addAction("(keine weiteren verfügbar)")
            no_cols.setEnabled(False)

        menu.exec_(header.mapToGlobal(pos))

    def _get_addable_columns(self):
        """Get columns that can be added (not currently selected in tree)"""
        available = {}
        all_cols = {**BASIC_COLUMNS, **TECHNICAL_COLUMNS}

        # Get currently shown column display names
        current_cols = set()
        for col in range(self.preview_table.columnCount()):
            item = self.preview_table.horizontalHeaderItem(col)
            if item:
                current_cols.add(item.text())

        # Find columns in tree that are unchecked
        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            if class_item.checkState(0) == QtCore.Qt.Unchecked:
                continue  # Skip unchecked classes

            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                if col_item.checkState(0) == QtCore.Qt.Unchecked:
                    col_data = col_item.data(0, QtCore.Qt.UserRole)
                    if col_data and col_data[0] == 'column':
                        col_key = col_data[1]
                        display_name = col_item.text(0)
                        if display_name not in current_cols:
                            available[col_key] = display_name

        return available

    def _remove_column_from_tree(self, col_display_name):
        """Remove a column by unchecking it in the tree"""
        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                if col_item.text(0) == col_display_name:
                    col_item.setCheckState(0, QtCore.Qt.Unchecked)
        # Preview will update via signal

    def _add_column_from_menu(self, col_key):
        """Add a column by checking it in the tree"""
        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            if class_item.checkState(0) == QtCore.Qt.Unchecked:
                continue

            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                col_data = col_item.data(0, QtCore.Qt.UserRole)
                if col_data and col_data[0] == 'column' and col_data[1] == col_key:
                    col_item.setCheckState(0, QtCore.Qt.Checked)
        # Preview will update via signal

    def _rename_column(self, original_name, col_index):
        """Rename a column for export only"""
        current_name = self.column_renames.get(original_name, original_name)
        new_name, ok = QtWidgets.QInputDialog.getText(
            self,
            "Spalte umbenennen",
            f"Neuer Name für '{original_name}':",
            QtWidgets.QLineEdit.Normal,
            current_name
        )
        if ok and new_name.strip():
            new_name = new_name.strip()
            if new_name != original_name:
                self.column_renames[original_name] = new_name
            elif original_name in self.column_renames:
                del self.column_renames[original_name]
            # Update header in preview
            header_item = self.preview_table.horizontalHeaderItem(col_index)
            if header_item:
                header_item.setText(new_name)
                header_item.setToolTip(f"Original: {original_name}")

    def _update_excluded_column_display(self):
        """Update visual display of excluded columns"""
        for col in range(self.preview_table.columnCount()):
            header_item = self.preview_table.horizontalHeaderItem(col)
            if not header_item:
                continue

            col_name = header_item.text()
            is_excluded = col_name in self.excluded_columns

            # Update header appearance
            if is_excluded:
                header_item.setForeground(QtGui.QColor(150, 150, 150))
                font = header_item.font()
                font.setStrikeOut(True)
                header_item.setFont(font)
            else:
                # Reset to default
                header_item.setForeground(QtGui.QColor(255, 255, 255) if self.palette().window().color().lightness() < 128 else QtGui.QColor(0, 0, 0))
                font = header_item.font()
                font.setStrikeOut(False)
                header_item.setFont(font)

            # Update all cells in the column
            for row in range(self.preview_table.rowCount()):
                item = self.preview_table.item(row, col)
                if item:
                    if is_excluded:
                        item.setForeground(QtGui.QColor(150, 150, 150))
                        font = item.font()
                        font.setStrikeOut(True)
                        item.setFont(font)
                    else:
                        # Check if row is excluded
                        row_excluded = row in self.excluded_row_indices
                        if row_excluded:
                            item.setForeground(QtGui.QColor(150, 150, 150))
                            font = item.font()
                            font.setStrikeOut(True)
                            item.setFont(font)
                        else:
                            item.setForeground(QtGui.QColor(255, 255, 255) if self.palette().window().color().lightness() < 128 else QtGui.QColor(0, 0, 0))
                            font = item.font()
                            font.setStrikeOut(False)
                            item.setFont(font)

        # Update excluded columns label
        excluded_col_count = len(self.excluded_columns)
        excluded_row_count = len(self.excluded_row_indices)
        parts = []
        if excluded_row_count > 0:
            parts.append(f"{excluded_row_count} Zeilen")
        if excluded_col_count > 0:
            parts.append(f"{excluded_col_count} Spalten")
        if parts:
            self.excluded_label.setText(f"(ausgeschlossen: {', '.join(parts)})")
        else:
            self.excluded_label.setText("")

    def _include_all_columns(self):
        """Include all columns back in export"""
        self.excluded_columns.clear()
        self._update_excluded_column_display()
        self._update_export_summary()

    def _exclude_selected_rows(self):
        """Exclude selected rows from export"""
        selected_rows = self.preview_table.selectionModel().selectedRows()
        for idx in selected_rows:
            self.excluded_row_indices.add(idx.row())

        # Update visual - gray out excluded rows
        self._update_excluded_row_display()
        self.preview_table.clearSelection()

    def _include_all_rows(self):
        """Include all rows back in export"""
        self.excluded_row_indices.clear()
        self._update_excluded_row_display()

    def _include_all(self):
        """Include all rows and columns back in export"""
        self.excluded_row_indices.clear()
        self.excluded_columns.clear()
        self._update_excluded_row_display()
        self._update_excluded_column_display()

    def _update_excluded_row_display(self):
        """Update the visual display of excluded rows"""
        for row in range(self.preview_table.rowCount()):
            is_excluded = row in self.excluded_row_indices
            for col in range(self.preview_table.columnCount()):
                item = self.preview_table.item(row, col)
                if item:
                    if is_excluded:
                        item.setForeground(QtGui.QColor(150, 150, 150))
                        font = item.font()
                        font.setStrikeOut(True)
                        item.setFont(font)
                    else:
                        item.setForeground(QtGui.QColor(255, 255, 255) if self.palette().window().color().lightness() < 128 else QtGui.QColor(0, 0, 0))
                        font = item.font()
                        font.setStrikeOut(False)
                        item.setFont(font)

        # Update excluded count
        excluded_count = len(self.excluded_row_indices)
        if excluded_count > 0:
            self.excluded_label.setText(f"({excluded_count} ausgeschlossen)")
        else:
            self.excluded_label.setText("")

        # Update summary
        self._update_export_summary()

    def _update_export_summary(self):
        """Update the export summary with excluded rows and columns"""
        total_rows = self.preview_table.rowCount()
        excluded_rows = len(self.excluded_row_indices)
        exported_rows = total_rows - excluded_rows

        # Count columns (minus Klasse reference column and excluded columns)
        total_cols = max(0, self.preview_table.columnCount() - 1)  # Minus Klasse column
        excluded_cols = len(self.excluded_columns)
        exported_cols = total_cols - excluded_cols

        format_text = self.format_combo.currentText().split()[0]
        self.summary_label.setText(f"{exported_rows} Zeilen × {exported_cols} Spalten → {format_text}")

    def _update_preview(self):
        """Update the data preview table and statistics"""
        try:
            config = self.get_export_config()

            if not config['classes']:
                self.preview_table.setRowCount(0)
                self.preview_table.setColumnCount(0)
                self.preview_label.setText("Keine Klassen ausgewählt")
                self.row_count_label.setText("")
                self.stats_label.setText("Keine Daten ausgewählt")
                self.summary_label.setText("Wählen Sie Daten zum Exportieren")
                return

            df_base = self.workspace.df_all.copy()
            if df_base.empty:
                self.preview_label.setText("Keine Daten")
                return

            # Collect data and statistics
            all_rows = []
            all_columns = []
            display_names = {}
            class_counts = {}

            for cls_name, cls_config in config['classes'].items():
                df_class = df_base[df_base['cls'] == cls_name]
                if df_class.empty:
                    continue

                class_counts[cls_name] = len(df_class)
                cols = cls_config['columns']
                names = cls_config['display_names']
                existing_cols = [c for c in cols if c in df_class.columns]

                for col in existing_cols:
                    if col not in all_columns:
                        all_columns.append(col)
                        display_names[col] = names.get(col, col)

                for _, row in df_class.iterrows():
                    row_data = {'_class': cls_name}
                    for col in existing_cols:
                        val = row.get(col, '')
                        if hasattr(val, '__iter__') and not isinstance(val, str):
                            val = str(list(val)) if len(val) > 0 else ''
                        elif pd.isna(val):
                            val = ''
                        row_data[col] = val
                    all_rows.append(row_data)

            if not all_rows:
                self.preview_label.setText("Keine Daten nach Filterung")
                self.preview_table.setRowCount(0)
                return

            # Limit preview
            max_rows = 200
            preview_rows = all_rows[:max_rows]
            total = len(all_rows)

            # Setup table - Klasse is reference only (not exported)
            headers = ['Klasse (Referenz)'] + [display_names.get(c, c) for c in all_columns]
            self.preview_table.setColumnCount(len(headers))
            self.preview_table.setRowCount(len(preview_rows))
            self.preview_table.setHorizontalHeaderLabels(headers)

            # Populate - gray out Klasse column
            for row_idx, row_data in enumerate(preview_rows):
                # Klasse column - grayed out (reference only)
                cls_item = QtWidgets.QTableWidgetItem(row_data.get('_class', ''))
                cls_item.setForeground(QtGui.QColor(128, 128, 128))  # Gray text
                font = cls_item.font()
                font.setItalic(True)
                cls_item.setFont(font)
                self.preview_table.setItem(row_idx, 0, cls_item)

                # Data columns
                for col_idx, col in enumerate(all_columns):
                    item = QtWidgets.QTableWidgetItem(str(row_data.get(col, '')))
                    self.preview_table.setItem(row_idx, col_idx + 1, item)

            self.preview_table.resizeColumnsToContents()
            self.preview_label.setText("<b>Vorschau:</b> Klicken Sie auf Zeilen um diese auszuwählen")

            if total > max_rows:
                self.row_count_label.setText(f"{total} Zeilen (zeige {max_rows})")
            else:
                self.row_count_label.setText(f"{total} Zeilen")

            # Clear excluded rows and columns when data changes
            self.excluded_row_indices.clear()
            self.excluded_columns.clear()
            self.excluded_label.setText("")

            # Update statistics
            stats_text = " | ".join([f"{name}: {count}" for name, count in class_counts.items()])
            self.stats_label.setText(stats_text if stats_text else "Keine Daten")

            # Update summary
            self._update_export_summary()

            # Update position hint
            self._update_position_hint()

        except Exception as e:
            self.preview_label.setText(f"Fehler: {e}")
            print(f"Preview error: {e}")

    def _on_target_changed(self):
        """Handle target radio button change"""
        is_existing = self.radio_existing.isChecked()
        self.existing_file_edit.setEnabled(is_existing)
        self.btn_browse.setEnabled(is_existing)
        self.position_widget.setVisible(is_existing)
        self.sheet_combo.setEnabled(is_existing and self.sheet_combo.count() > 0)
        self.pos_column.setEnabled(is_existing)
        self.pos_row.setEnabled(is_existing and not self.check_append.isChecked())
        self.check_append.setEnabled(is_existing)
        self._update_position_hint()

    def _update_position_hint(self):
        """Update the position hint label and visual preview"""
        if not self.radio_existing.isChecked():
            self.position_hint.setText("")
            self.excel_preview.setText("")
            self._cached_scan_result = None
            return

        col = self.pos_column.text().upper() or "A"
        row = self.pos_row.value()

        # Calculate column count from preview
        col_count = self.preview_table.columnCount()
        row_count = self.preview_table.rowCount()
        include_header = self.check_include_header.isChecked()
        total_rows = row_count + (1 if include_header else 0)

        if col_count > 0:
            try:
                start_col_idx = column_index_from_string(col)
                end_col_idx = start_col_idx + col_count - 1
                end_col = get_column_letter(end_col_idx)

                # Get file path for scanning
                file_path = self.existing_file_edit.property('full_path')
                sheet_name = self.sheet_combo.currentText()

                # Scan target area for existing data/formulas
                scan_result = None
                warning_text = ""
                if file_path and sheet_name:
                    scan_result = self._scan_target_area(
                        file_path, sheet_name, col, row,
                        col_count, min(total_rows, 10)  # Scan up to 10 rows for preview
                    )
                    self._cached_scan_result = scan_result

                    if scan_result['formula_count'] > 0:
                        warning_text = f"  {scan_result['formula_count']} Formeln!"
                    elif scan_result['data_count'] > 0:
                        warning_text = f"  {scan_result['data_count']} Zellen mit Daten"

                self.position_hint.setText(
                    f"→ Daten: {col}{row} bis {end_col}{row + total_rows - 1}{warning_text}"
                )

                # Show preview with actual cell contents if available
                if scan_result and (scan_result['formula_count'] > 0 or scan_result['data_count'] > 0):
                    preview_text = self._get_target_preview_text(
                        scan_result, col, row, col_count, min(total_rows, 5)
                    )
                    self.excel_preview.setText(preview_text)
                else:
                    # Show simple preview for empty target area
                    preview_lines = []
                    show_cols = min(col_count, 4)
                    show_rows = min(total_rows, 5)

                    col_letters = [get_column_letter(start_col_idx + i) for i in range(show_cols)]
                    if col_count > show_cols:
                        col_letters.append("..")
                    header = "     " + "".join(f" {c:^4}" for c in col_letters)
                    preview_lines.append(header)

                    for r in range(show_rows):
                        row_num = row + r
                        cells = [" [██]" for _ in range(show_cols)]
                        if col_count > show_cols:
                            cells.append(" ..")
                        preview_lines.append(f"{row_num:>4} " + "".join(cells))

                    if total_rows > show_rows:
                        preview_lines.append("  ...")

                    preview_lines.append("")
                    preview_lines.append("Zielbereich ist leer")
                    self.excel_preview.setText("\n".join(preview_lines))
            except Exception as e:
                self.position_hint.setText(f"→ Ab {col}{row}")
                self.excel_preview.setText("")
                self._cached_scan_result = None
        else:
            self.position_hint.setText(f"→ Ab {col}{row}")
            self.excel_preview.setText("")
            self._cached_scan_result = None

    def _scan_target_area(self, file_path, sheet_name, start_col, start_row, num_cols, num_rows):
        """Scan target area in Excel file for existing data and formulas.

        Returns:
            dict with keys:
                'cells': list of cell info dicts with 'address', 'value', 'is_formula', 'formula'
                'formula_count': number of cells with formulas
                'data_count': number of cells with data (non-empty)
                'formulas': list of formula cell addresses
        """
        result = {
            'cells': [],
            'formula_count': 0,
            'data_count': 0,
            'formulas': [],
            'preview_grid': []  # 2D array for preview display
        }

        if not file_path or not os.path.exists(file_path):
            return result

        wb = None
        try:
            wb = load_workbook(file_path, read_only=False, data_only=False)
            if sheet_name not in wb.sheetnames:
                return result

            ws = wb[sheet_name]
            start_col_idx = column_index_from_string(start_col)

            for r in range(num_rows):
                row_cells = []
                for c in range(num_cols):
                    cell = ws.cell(row=start_row + r, column=start_col_idx + c)
                    cell_addr = f"{get_column_letter(start_col_idx + c)}{start_row + r}"

                    cell_info = {
                        'address': cell_addr,
                        'value': cell.value,
                        'is_formula': False,
                        'formula': None
                    }

                    # Check if cell contains a formula
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_info['is_formula'] = True
                        cell_info['formula'] = cell.value
                        result['formula_count'] += 1
                        result['formulas'].append(cell_addr)
                    elif cell.value is not None and cell.value != '':
                        result['data_count'] += 1

                    result['cells'].append(cell_info)
                    row_cells.append(cell_info)
                result['preview_grid'].append(row_cells)

        except Exception as e:
            print(f"Error scanning target area: {e}")
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass

        return result

    def _get_target_preview_text(self, scan_result, start_col, start_row, num_cols, num_rows):
        """Generate preview text showing target cells with their current content."""
        if not scan_result or not scan_result.get('preview_grid'):
            return ""

        lines = []
        start_col_idx = column_index_from_string(start_col)
        show_cols = min(num_cols, 4)
        show_rows = min(num_rows, 5)

        # Header row with column letters
        col_letters = [get_column_letter(start_col_idx + i) for i in range(show_cols)]
        if num_cols > show_cols:
            col_letters.append("..")
        header = "     " + "".join(f" {c:^4}" for c in col_letters)
        lines.append(header)

        # Data rows
        grid = scan_result['preview_grid']
        for r in range(min(show_rows, len(grid))):
            row_num = start_row + r
            cells = []
            for c in range(min(show_cols, len(grid[r]))):
                cell_info = grid[r][c]
                if cell_info['is_formula']:
                    cells.append(" [f] ")  # Formula indicator
                elif cell_info['value'] is not None and cell_info['value'] != '':
                    # Truncate value for display
                    val_str = str(cell_info['value'])[:3]
                    cells.append(f"[{val_str:^3}]")
                else:
                    cells.append(" [ ] ")  # Empty cell
            if num_cols > show_cols:
                cells.append(" ..")
            lines.append(f"{row_num:>4} " + "".join(cells))

        if num_rows > show_rows:
            lines.append("  ...")

        # Add legend
        if scan_result['formula_count'] > 0 or scan_result['data_count'] > 0:
            lines.append("")
            legend = []
            if scan_result['formula_count'] > 0:
                legend.append(f"[f]={scan_result['formula_count']} Formeln")
            if scan_result['data_count'] > 0:
                legend.append(f"[x]={scan_result['data_count']} Daten")
            lines.append(" ".join(legend))

        return "\n".join(lines)

    def _count_formulas_outside_target(self, file_path, sheet_name, start_col, start_row, num_cols, num_rows):
        """Count formulas in the sheet that are outside the target area."""
        formula_count = 0
        formula_cells = []

        if not file_path or not os.path.exists(file_path):
            return {'count': 0, 'cells': []}

        wb = None
        try:
            wb = load_workbook(file_path, read_only=False, data_only=False)
            if sheet_name not in wb.sheetnames:
                return {'count': 0, 'cells': []}

            ws = wb[sheet_name]
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = start_col_idx + num_cols - 1
            end_row = start_row + num_rows - 1

            # Scan all cells in used range
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Check if cell is outside target area
                        cell_col_idx = cell.column
                        cell_row = cell.row

                        is_outside = (
                            cell_row < start_row or cell_row > end_row or
                            cell_col_idx < start_col_idx or cell_col_idx > end_col_idx
                        )

                        if is_outside:
                            formula_count += 1
                            formula_cells.append(f"{get_column_letter(cell_col_idx)}{cell_row}")
        except Exception as e:
            print(f"Error counting formulas: {e}")
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass

        return {'count': formula_count, 'cells': formula_cells}

    def _verify_structure_after_export(self, file_path, sheet_name, start_col, start_row, num_cols, num_rows, formulas_before):
        """Verify that formulas outside target area are preserved after export."""
        result = {
            'success': True,
            'formulas_preserved': 0,
            'formulas_lost': 0,
            'message': ''
        }

        try:
            formulas_after = self._count_formulas_outside_target(
                file_path, sheet_name, start_col, start_row, num_cols, num_rows
            )

            result['formulas_preserved'] = formulas_after['count']

            if formulas_before['count'] != formulas_after['count']:
                lost = formulas_before['count'] - formulas_after['count']
                result['success'] = False
                result['formulas_lost'] = lost
                result['message'] = (
                    f"{lost} Formeln außerhalb des Zielbereichs wurden verändert oder gelöscht.\n"
                    "Bitte überprüfen Sie die Datei manuell."
                )
            else:
                result['message'] = f"Alle {formulas_after['count']} Formeln außerhalb des Zielbereichs erhalten."

        except Exception as e:
            result['success'] = False
            result['message'] = f"Strukturprüfung fehlgeschlagen: {e}"

        return result

    def _find_safe_position(self, file_path, sheet_name, start_col, num_cols, num_rows, current_row=1):
        """Find the next row where export won't overwrite formulas or data.

        Returns:
            dict with 'row' (safe row number) and 'reason' (why this row was chosen)
        """
        if not file_path or not os.path.exists(file_path):
            return {'row': current_row, 'reason': 'Datei nicht gefunden'}

        wb = None
        result = {'row': current_row, 'reason': 'Unbekannt'}
        try:
            wb = load_workbook(file_path, read_only=False, data_only=False)
            if sheet_name not in wb.sheetnames:
                return {'row': current_row, 'reason': 'Blatt nicht gefunden'}

            ws = wb[sheet_name]
            start_col_idx = column_index_from_string(start_col)
            max_row = ws.max_row or 1

            # Search from current row downward for a safe position
            test_row = current_row
            max_search = max_row + 100  # Don't search forever

            while test_row <= max_search:
                # Check if this row range is safe (no formulas or data)
                is_safe = True
                for r in range(num_rows):
                    for c in range(num_cols):
                        cell = ws.cell(row=test_row + r, column=start_col_idx + c)
                        if cell.value is not None and cell.value != '':
                            is_safe = False
                            break
                    if not is_safe:
                        break

                if is_safe:
                    if test_row == current_row:
                        result = {'row': test_row, 'reason': 'Aktuelle Position ist sicher'}
                    elif test_row > max_row:
                        result = {'row': test_row, 'reason': f'Nach letzter Zeile ({max_row})'}
                    else:
                        result = {'row': test_row, 'reason': f'Nächster freier Bereich ab Zeile {test_row}'}
                    return result

                # Try next row
                test_row += 1

            # If no safe position found, suggest after max_row
            result = {'row': max_row + 1, 'reason': f'Nach letzter Zeile ({max_row})'}
            return result

        except Exception as e:
            print(f"Error finding safe position: {e}")
            return {'row': current_row, 'reason': f'Fehler: {e}'}
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass

    def _browse_file(self):
        """Browse for existing Excel file"""
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Excel-Datei auswählen", "", "Excel (*.xlsx *.xls)"
        )
        if not path:
            return

        self.existing_file_edit.setText(os.path.basename(path))
        self.existing_file_edit.setToolTip(path)
        self.existing_file_edit.setProperty('full_path', path)

        try:
            wb = load_workbook(path, read_only=True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(wb.sheetnames)
            self.sheet_combo.setEnabled(True)
            wb.close()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Fehler", f"Datei nicht lesbar: {e}")

    def get_export_config(self) -> dict:
        """Get export configuration from tree"""
        # Get CSV separator
        csv_sep_map = {
            "Semikolon (;)": ";",
            "Komma (,)": ",",
            "Tab": "\t",
            "Pipe (|)": "|"
        }
        csv_sep = csv_sep_map.get(self.csv_separator_combo.currentText(), ";")

        # Get empty cell replacement
        empty_cell_map = {
            "Leer lassen": "",
            "Mit '-' füllen": "-",
            "Mit 'N/A' füllen": "N/A"
        }
        empty_cell_value = empty_cell_map.get(self.empty_cell_combo.currentText(), "")

        # Get column order from preview table (user may have dragged columns)
        column_order = []
        header = self.preview_table.horizontalHeader()
        for visual_idx in range(self.preview_table.columnCount()):
            logical_idx = header.logicalIndex(visual_idx)
            item = self.preview_table.horizontalHeaderItem(logical_idx)
            if item:
                column_order.append(item.text())

        config = {
            'classes': {},
            'options': {
                'apply_filters': self.check_apply_filters.isChecked(),
                'auto_width': self.check_auto_width.isChecked(),
                'freeze_header': self.check_freeze_header.isChecked(),
                'sort_by': self.sort_combo.currentText(),
                'include_header': self.check_include_header.isChecked(),
                'include_empty': True,
                'format': self.format_combo.currentText(),
                'excluded_rows': list(self.excluded_row_indices),
                'excluded_columns': set(self.excluded_columns),
                'column_renames': dict(self.column_renames),
                'column_order': column_order,
                'grouping': self.grouping_combo.currentText(),
                'empty_cell_value': empty_cell_value,
                'number_format': self.number_format_combo.currentText(),
                'coord_scale': self.coord_scale_combo.currentText(),
                'csv_separator': csv_sep,
                'json_orient': self.json_orient_combo.currentText(),
            },
            'target': self._get_target_config()
        }

        # Extract from tree
        for i in range(self.tree.topLevelItemCount()):
            class_item = self.tree.topLevelItem(i)
            if class_item.checkState(0) == QtCore.Qt.Unchecked:
                continue

            data = class_item.data(0, QtCore.Qt.UserRole)
            if not data or data[0] != 'class':
                continue

            cls_name = data[1]
            columns = []
            display_names = {}

            for j in range(class_item.childCount()):
                col_item = class_item.child(j)
                if col_item.checkState(0) != QtCore.Qt.Checked:
                    continue

                col_data = col_item.data(0, QtCore.Qt.UserRole)
                if col_data and col_data[0] == 'column':
                    col_key = col_data[1]
                    columns.append(col_key)
                    display_names[col_key] = col_item.text(0)

            if columns:
                config['classes'][cls_name] = {
                    'columns': columns,
                    'display_names': display_names
                }

        return config

    def _get_target_config(self) -> dict:
        """Get target configuration"""
        if self.radio_new.isChecked():
            return {'mode': 'new_file'}

        full_path = self.existing_file_edit.property('full_path') or ''
        return {
            'mode': 'existing_file',
            'file_path': full_path,
            'sheet_name': self.sheet_combo.currentText(),
            'start_column': self.pos_column.text().upper() or 'A',
            'start_row': self.pos_row.value(),
            'smart_append': self.check_append.isChecked(),
            'include_header': self.check_include_header.isChecked(),
        }

    def _on_export(self):
        """Handle export button - exports and stays open"""
        config = self.get_export_config()

        if not config['classes']:
            QtWidgets.QMessageBox.warning(self, "Fehler", "Bitte wählen Sie mindestens eine Klasse aus.")
            return

        target = config['target']
        if target['mode'] == 'existing_file':
            if not target['file_path'] or not os.path.exists(target['file_path']):
                QtWidgets.QMessageBox.warning(self, "Fehler", "Bitte wählen Sie eine gültige Datei aus.")
                return

            # Validate sheet name
            if not target.get('sheet_name'):
                QtWidgets.QMessageBox.warning(self, "Fehler", "Bitte wählen Sie ein Blatt aus.")
                return

            # Validate column letter
            col_str = target.get('start_column', 'A')
            if not col_str or not col_str.isalpha():
                QtWidgets.QMessageBox.warning(
                    self, "Fehler",
                    f"Ungültige Spalte: '{col_str}'\n\nBitte geben Sie einen gültigen Spaltenbuchstaben ein (z.B. A, B, AA)."
                )
                return

        # Store config for parent to access
        self._export_config = config

        # Try to export directly
        try:
            if target['mode'] == 'new_file':
                self._do_export_new_file(config)
            else:
                self._do_export_existing_file(config)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export Fehler", f"Fehler beim Export:\n{str(e)}")

    def _do_export_new_file(self, config):
        """Export to a new file (Excel, CSV, or JSON)"""
        format_type = config['options'].get('format', 'Excel (.xlsx)')
        is_csv = 'CSV' in format_type
        is_json = 'JSON' in format_type

        # Ask for save location based on format
        if is_csv:
            file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "CSV speichern", "", "CSV (*.csv)"
            )
            if file_path and not file_path.endswith('.csv'):
                file_path += '.csv'
        elif is_json:
            file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "JSON speichern", "", "JSON (*.json)"
            )
            if file_path and not file_path.endswith('.json'):
                file_path += '.json'
        else:
            file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Excel speichern", "", "Excel (*.xlsx)"
            )
            if file_path and not file_path.endswith('.xlsx'):
                file_path += '.xlsx'

        if not file_path:
            return

        grouping = config['options'].get('grouping', 'Keine')

        # Handle separate sheets export (Excel only)
        if 'separate Blätter' in grouping and not is_csv and not is_json:
            self._export_separate_sheets(config, file_path)
            return

        # Handle JSON export
        if is_json:
            self._export_json(config, file_path)
            return

        # Build single dataframe
        df = self._build_export_dataframe(config)
        if df.empty:
            QtWidgets.QMessageBox.warning(self, "Keine Daten", "Keine Daten zum Exportieren.")
            return

        # Export based on format
        try:
            if is_csv:
                csv_sep = config['options'].get('csv_separator', ';')
                df.to_csv(file_path, index=False, sep=csv_sep, encoding='utf-8-sig')
                QtWidgets.QMessageBox.information(self, "Erfolg", f"CSV exportiert nach:\n{file_path}")
            else:
                df.to_excel(file_path, index=False, sheet_name='Export')
                QtWidgets.QMessageBox.information(self, "Erfolg", f"Excel exportiert nach:\n{file_path}")
        except PermissionError:
            QtWidgets.QMessageBox.critical(
                self, "Datei gesperrt",
                f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
                "Die Datei ist wahrscheinlich geöffnet.\n"
                "Bitte schließen Sie die Datei und versuchen Sie es erneut."
            )

    def _export_separate_sheets(self, config, file_path):
        """Export each class to a separate sheet"""
        from openpyxl import Workbook

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        df_base = self.workspace.df_all.copy()
        sheets_created = 0

        for cls_name, cls_config in config['classes'].items():
            df_class = df_base[df_base['cls'] == cls_name]
            if df_class.empty:
                continue

            # Build dataframe for this class
            class_config = {
                'classes': {cls_name: cls_config},
                'options': config['options']
            }
            df = self._build_export_dataframe(class_config)

            if df.empty:
                continue

            # Create sheet (sheet name max 31 chars)
            sheet_name = cls_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Write header
            if config['options'].get('include_header', True):
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)

            # Write data
            data_start_row = 2 if config['options'].get('include_header', True) else 1
            for row_num, (_, row) in enumerate(df.iterrows()):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=data_start_row + row_num, column=col_idx, value=value)

            sheets_created += 1

        if sheets_created == 0:
            QtWidgets.QMessageBox.warning(self, "Keine Daten", "Keine Daten zum Exportieren.")
            return

        try:
            wb.save(file_path)
            QtWidgets.QMessageBox.information(
                self, "Erfolg",
                f"Excel exportiert nach:\n{file_path}\n\n{sheets_created} Blätter erstellt."
            )
        except PermissionError:
            QtWidgets.QMessageBox.critical(
                self, "Datei gesperrt",
                f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
                "Die Datei ist wahrscheinlich in Excel geöffnet.\n"
                "Bitte schließen Sie die Datei in Excel und versuchen Sie es erneut."
            )

    def _export_json(self, config, file_path):
        """Export data to JSON file"""
        import json

        json_orient = config['options'].get('json_orient', 'Records (Liste von Objekten)')

        # Determine export structure
        if 'Nach Klasse gruppiert' in json_orient:
            # Export grouped by class
            export_data = {}
            df_base = self.workspace.df_all.copy()

            for cls_name, cls_config in config['classes'].items():
                # Build dataframe for this class
                class_config = {
                    'classes': {cls_name: cls_config},
                    'options': config['options']
                }
                df = self._build_export_dataframe(class_config)

                if not df.empty:
                    # Convert to list of dicts (records)
                    export_data[cls_name] = df.to_dict(orient='records')

            if not export_data:
                QtWidgets.QMessageBox.warning(self, "Keine Daten", "Keine Daten zum Exportieren.")
                return

        else:
            # Build single dataframe
            df = self._build_export_dataframe(config)
            if df.empty:
                QtWidgets.QMessageBox.warning(self, "Keine Daten", "Keine Daten zum Exportieren.")
                return

            # Convert based on orient option
            if 'Records' in json_orient:
                export_data = df.to_dict(orient='records')
            elif 'Tabelle' in json_orient:
                export_data = {
                    'columns': list(df.columns),
                    'data': df.values.tolist()
                }
            elif 'Index' in json_orient:
                export_data = df.to_dict(orient='index')
            else:
                export_data = df.to_dict(orient='records')

        # Write JSON file
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
            QtWidgets.QMessageBox.information(
                self, "Erfolg",
                f"JSON exportiert nach:\n{file_path}"
            )
        except PermissionError:
            QtWidgets.QMessageBox.critical(
                self, "Datei gesperrt",
                f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
                "Die Datei ist möglicherweise geöffnet.\n"
                "Bitte schließen Sie die Datei und versuchen Sie es erneut."
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Fehler", f"JSON Export fehlgeschlagen:\n{str(e)}")

    def _do_export_existing_file(self, config):
        """Export to an existing Excel file"""
        target = config['target']
        file_path = target['file_path']
        sheet_name = target['sheet_name']
        start_col = target['start_column']
        start_row = target['start_row']

        # Build dataframe
        df = self._build_export_dataframe(config)
        if df.empty:
            QtWidgets.QMessageBox.warning(self, "Keine Daten", "Keine Daten zum Exportieren.")
            return

        include_header = config['options'].get('include_header', True)
        total_rows = len(df) + (1 if include_header else 0)
        total_cols = len(df.columns)

        # Check for formulas/data in target area BEFORE export
        scan_result = self._scan_target_area(
            file_path, sheet_name, start_col, start_row, total_cols, total_rows
        )

        # Track if we should skip formula cells during write
        skip_formula_cells = False

        # Warn user if formulas or data will be overwritten
        if scan_result['formula_count'] > 0 or scan_result['data_count'] > 0:
            # Find a safe alternative position
            safe_pos = self._find_safe_position(
                file_path, sheet_name, start_col, total_cols, total_rows, start_row
            )

            # Build warning message
            if scan_result['formula_count'] > 0:
                formula_list = ", ".join(scan_result['formulas'][:5])
                if len(scan_result['formulas']) > 5:
                    formula_list += f" ..."
                warning_title = " Formeln im Zielbereich!"
                warning_text = (
                    f"Im Zielbereich ({start_col}{start_row}) befinden sich "
                    f"{scan_result['formula_count']} Formeln:\n{formula_list}\n\n"
                )
            else:
                warning_title = "Daten im Zielbereich"
                warning_text = (
                    f"Im Zielbereich ({start_col}{start_row}) befinden sich "
                    f"{scan_result['data_count']} Zellen mit Daten.\n\n"
                )

            # Create custom dialog with options
            msg_box = QtWidgets.QMessageBox(self)
            msg_box.setWindowTitle(warning_title)
            msg_box.setIcon(QtWidgets.QMessageBox.Warning if scan_result['formula_count'] > 0 else QtWidgets.QMessageBox.Question)

            # Add options
            warning_text += "Was möchten Sie tun?"
            msg_box.setText(warning_text)

            # Option to keep formulas (skip formula cells during export)
            if scan_result['formula_count'] > 0:
                btn_keep_formulas = msg_box.addButton(" Formeln behalten", QtWidgets.QMessageBox.ActionRole)
                btn_keep_formulas.setToolTip("Daten exportieren, aber Formelzellen überspringen")
            else:
                btn_keep_formulas = None

            # Option to auto-adjust to safe position
            if safe_pos['row'] != start_row:
                btn_auto = msg_box.addButton(f"↓ Zeile {safe_pos['row']}", QtWidgets.QMessageBox.ActionRole)
            else:
                btn_auto = None

            btn_adjust = msg_box.addButton(" Anpassen", QtWidgets.QMessageBox.ActionRole)
            btn_override = msg_box.addButton(" Überschreiben", QtWidgets.QMessageBox.DestructiveRole)
            btn_cancel = msg_box.addButton("Abbrechen", QtWidgets.QMessageBox.RejectRole)

            msg_box.exec_()
            clicked = msg_box.clickedButton()

            if clicked == btn_keep_formulas:
                # Export but skip formula cells
                skip_formula_cells = True
            elif clicked == btn_auto:
                # Use the safe position
                start_row = safe_pos['row']
            elif clicked == btn_adjust:
                # User wants to adjust manually - close dialog and return
                QtWidgets.QMessageBox.information(
                    self, "Position anpassen",
                    f"Passen Sie die Startzeile im Feld 'Zeile' an.\n\n"
                    f"Empfehlung: Zeile {safe_pos['row']} oder höher."
                )
                return
            elif clicked == btn_override:
                # User wants to override - continue with original position
                pass
            else:
                # Cancel
                return

        # Count formulas outside target area BEFORE export (for structure verification)
        formulas_before = self._count_formulas_outside_target(
            file_path, sheet_name, start_col, start_row, total_cols, total_rows
        )

        # Load existing workbook
        from openpyxl import load_workbook
        try:
            wb = load_workbook(file_path)
        except PermissionError:
            QtWidgets.QMessageBox.critical(
                self, "Datei gesperrt",
                f"Die Datei kann nicht geöffnet werden:\n{file_path}\n\n"
                "Die Datei ist wahrscheinlich in Excel geöffnet.\n"
                "Bitte schließen Sie die Datei in Excel und versuchen Sie es erneut."
            )
            return

        if sheet_name not in wb.sheetnames:
            QtWidgets.QMessageBox.warning(self, "Fehler", f"Blatt '{sheet_name}' nicht gefunden.")
            return

        ws = wb[sheet_name]

        # Smart append - find last row
        if target.get('smart_append', False):
            start_row = ws.max_row + 1

        # Write data
        start_col_idx = column_index_from_string(start_col)
        skipped_cells = 0  # Count skipped formula cells

        row_offset = 0
        if include_header:
            for col_idx, col_name in enumerate(df.columns):
                cell = ws.cell(row=start_row, column=start_col_idx + col_idx)
                # Skip merged cells (read-only)
                if isinstance(cell, MergedCell):
                    continue
                # Skip if cell has formula and skip_formula_cells is enabled
                if skip_formula_cells and cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    skipped_cells += 1
                    continue
                cell.value = col_name
            row_offset = 1

        for row_num, (_, row) in enumerate(df.iterrows()):
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + row_offset + row_num, column=start_col_idx + col_idx)
                # Skip merged cells (read-only)
                if isinstance(cell, MergedCell):
                    continue
                # Skip if cell has formula and skip_formula_cells is enabled
                if skip_formula_cells and cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    skipped_cells += 1
                    continue
                cell.value = value

        try:
            wb.save(file_path)
        except PermissionError:
            QtWidgets.QMessageBox.critical(
                self, "Datei gesperrt",
                f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
                "Die Datei ist wahrscheinlich in Excel geöffnet.\n"
                "Bitte schließen Sie die Datei in Excel und versuchen Sie es erneut."
            )
            return

        # Verify structure after export
        verification_result = self._verify_structure_after_export(
            file_path, sheet_name, start_col, start_row, total_cols, total_rows, formulas_before
        )

        # Show success message with verification info
        skip_info = ""
        if skipped_cells > 0:
            skip_info = f"  • {skipped_cells} Formelzellen übersprungen (Formeln erhalten)\n"

        if verification_result['success']:
            QtWidgets.QMessageBox.information(
                self, " Export erfolgreich",
                f"Daten exportiert nach:\n{file_path}\nBlatt: {sheet_name}\n\n"
                f" Strukturprüfung bestanden\n"
                f"{skip_info}"
                f"  • {verification_result['formulas_preserved']} Formeln außerhalb des Zielbereichs erhalten"
            )
        else:
            QtWidgets.QMessageBox.warning(
                self, " Export mit Warnung",
                f"Daten exportiert nach:\n{file_path}\nBlatt: {sheet_name}\n\n"
                f"{skip_info}"
                f" Strukturprüfung:\n{verification_result['message']}"
            )

    def _build_export_dataframe(self, config):
        """Build the export dataframe from config"""
        df_base = self.workspace.df_all.copy()
        excluded_rows = set(config['options'].get('excluded_rows', []))
        excluded_cols = config['options'].get('excluded_columns', set())

        all_rows = []
        row_counter = 0

        for cls_name, cls_config in config['classes'].items():
            df_class = df_base[df_base['cls'] == cls_name]
            if df_class.empty:
                continue

            cols = cls_config['columns']
            existing_cols = [c for c in cols if c in df_class.columns]

            for _, row in df_class.iterrows():
                # Skip excluded rows
                if row_counter in excluded_rows:
                    row_counter += 1
                    continue

                row_data = {}
                for col in existing_cols:
                    display_name = cls_config['display_names'].get(col, col)
                    # Skip excluded columns (by display name)
                    if display_name in excluded_cols:
                        continue
                    val = row.get(col, '')
                    if hasattr(val, '__iter__') and not isinstance(val, str):
                        val = str(list(val)) if len(val) > 0 else ''
                    elif pd.isna(val):
                        val = ''
                    row_data[display_name] = val
                all_rows.append(row_data)
                row_counter += 1

        if not all_rows:
            return pd.DataFrame()

        df = pd.DataFrame(all_rows)

        # Apply sorting
        sort_by = config['options'].get('sort_by', 'Keine')
        if sort_by and sort_by != 'Keine':
            try:
                if 'Klasse' in sort_by:
                    sort_col = 'Klasse (Referenz)' if 'Klasse (Referenz)' in df.columns else None
                    # Find any column with Klasse
                    if not sort_col:
                        for col in df.columns:
                            if 'klasse' in col.lower():
                                sort_col = col
                                break
                    ascending = '(A→Z)' in sort_by or '(1→9)' in sort_by
                    if sort_col and sort_col in df.columns:
                        df = df.sort_values(by=sort_col, ascending=ascending)
                elif 'Seite' in sort_by:
                    sort_col = None
                    for col in df.columns:
                        if 'seite' in col.lower() or col.lower() == 'page':
                            sort_col = col
                            break
                    ascending = '(1→9)' in sort_by
                    if sort_col and sort_col in df.columns:
                        df = df.sort_values(by=sort_col, ascending=ascending)
                elif 'Koordinate' in sort_by:
                    sort_col = None
                    for col in df.columns:
                        if 'koordinate' in col.lower() or 'coord' in col.lower():
                            sort_col = col
                            break
                    ascending = '(A→Z)' in sort_by
                    if sort_col and sort_col in df.columns:
                        df = df.sort_values(by=sort_col, ascending=ascending)
            except Exception as e:
                print(f"Sorting error: {e}")

        # Apply column renames
        column_renames = config['options'].get('column_renames', {})
        if column_renames:
            df = df.rename(columns=column_renames)

        # Apply coordinate scaling
        coord_scale = config['options'].get('coord_scale', 'Original')
        if coord_scale != 'Original':
            # Define which columns are coordinate columns
            coord_columns = ['x', 'y', 'width', 'height', 'x1', 'y1', 'x2', 'y2',
                           'center_x', 'center_y', 'X', 'Y', 'Breite', 'Höhe']

            # Determine scale factor
            scale_factors = {
                '×1000 (m→mm)': 1000,
                '×100 (m→cm)': 100,
                '×10': 10,
                '÷1000 (mm→m)': 0.001,
                '÷100 (cm→m)': 0.01
            }
            scale = scale_factors.get(coord_scale, 1)

            for col in df.columns:
                # Check if column name matches any coordinate column (case insensitive)
                col_lower = col.lower()
                is_coord = any(cc.lower() in col_lower for cc in coord_columns)

                if is_coord:
                    try:
                        numeric_col = pd.to_numeric(df[col], errors='coerce')
                        if numeric_col.notna().any():
                            df[col] = numeric_col * scale
                    except:
                        pass

        # Apply empty cell handling
        empty_cell_value = config['options'].get('empty_cell_value', '')
        if empty_cell_value:
            df = df.fillna(empty_cell_value)
            df = df.replace('', empty_cell_value)

        # Apply number formatting
        number_format = config['options'].get('number_format', 'Original')
        if number_format != 'Original':
            for col in df.columns:
                try:
                    # Try to convert to numeric and format
                    numeric_col = pd.to_numeric(df[col], errors='coerce')
                    if numeric_col.notna().any():
                        if number_format == '2 Dezimalen':
                            df[col] = numeric_col.apply(lambda x: f"{x:.2f}" if pd.notna(x) else df[col])
                        elif number_format == 'Ganzzahl':
                            df[col] = numeric_col.apply(lambda x: str(int(x)) if pd.notna(x) else df[col])
                except:
                    pass

        # Apply column order (from drag & drop)
        column_order = config['options'].get('column_order', [])
        if column_order:
            # Filter to only columns that exist in df
            ordered_cols = [c for c in column_order if c in df.columns]
            # Add any columns not in order
            remaining = [c for c in df.columns if c not in ordered_cols]
            df = df[ordered_cols + remaining]

        return df

    def _show_help(self):
        """Show comprehensive help dialog"""
        help_text = """
        <h2> Excel Export - Anleitung</h2>

        <hr>
        <h3> Schnellstart</h3>
        <ol>
            <li><b>Vorlagen wählen:</b> Klicken Sie auf <b> Techniker</b> oder <b> Entwickler</b></li>
            <li><b>Vorschau prüfen:</b> Die rechte Tabelle zeigt, was exportiert wird</li>
            <li><b>Exportieren:</b> Klicken Sie auf <b>Exportieren</b></li>
        </ol>

        <hr>
        <h3> Toolbar</h3>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;">
            <tr><td><b> Techniker</b></td><td>Nur Basis-Spalten (Text, Koordinaten, Fahrtrichtung)</td></tr>
            <tr><td><b> Entwickler</b></td><td>Alle Spalten inkl. technische Daten</td></tr>
            <tr><td><b> Alle</b></td><td>Alle Klassen und Spalten auswählen</td></tr>
            <tr><td><b> Keine</b></td><td>Alles abwählen</td></tr>
        </table>

        <hr>
        <h3> Datenvorschau</h3>
        <ul>
            <li><b>Spalten ziehen:</b> Ziehen Sie Spaltenköpfe um die Reihenfolge zu ändern</li>
            <li><b>Linksklick auf Spaltenkopf:</b> Spalte ein-/ausschließen</li>
            <li><b>Rechtsklick auf Spaltenkopf:</b> Kontextmenü mit mehr Optionen:
                <ul>
                    <li>Spalte umbenennen (nur für Export)</li>
                    <li>Spalte entfernen</li>
                    <li>Neue Spalte hinzufügen</li>
                </ul>
            </li>
            <li><b>Zeilen markieren + "Zeilen ausschließen":</b> Zeilen vom Export ausschließen</li>
        </ul>

        <hr>
        <h3> Optionen</h3>

        <h4>Basis-Optionen:</h4>
        <ul>
            <li><b>Mit Überschriften:</b> Spaltennamen in erster Zeile</li>
            <li><b>Spaltenbreite anpassen:</b> Automatische Spaltenbreite (nur Excel)</li>
            <li><b>Kopfzeile fixieren:</b> Header beim Scrollen sichtbar (nur Excel)</li>
            <li><b>Sortierung:</b> Aufsteigend (A→Z, 1→9) oder absteigend (Z→A, 9→1)</li>
        </ul>

        <h4>Erweiterte Optionen:</h4>
        <ul>
            <li><b>Gruppierung:</b>
                <ul>
                    <li><i>Keine</i> - Alle Daten in einer Tabelle</li>
                    <li><i>Nach Klasse (Abschnitte)</i> - Gruppiert nach Klasse</li>
                    <li><i>Nach Klasse (separate Blätter)</i> - Jede Klasse auf eigenem Excel-Blatt</li>
                </ul>
            </li>
            <li><b>Leere Zellen:</b> Leer lassen, mit '-' oder 'N/A' füllen</li>
            <li><b>Zahlen:</b> Original, 2 Dezimalen, oder Ganzzahl</li>
        </ul>

        <h4>CSV-Optionen (nur bei CSV-Export):</h4>
        <ul>
            <li><b>Semikolon (;)</b> - Empfohlen für deutsches Excel</li>
            <li><b>Komma (,)</b> - Internationaler Standard</li>
            <li><b>Tab / Pipe</b> - Alternative Trennzeichen</li>
        </ul>

        <hr>
        <h3> Export-Ziel</h3>
        <ul>
            <li><b>Neue Datei:</b> Erstellt eine neue Excel- oder CSV-Datei</li>
            <li><b>Bestehende:</b> Fügt Daten in vorhandene Excel-Datei ein
                <ul>
                    <li>Wählen Sie Blatt, Start-Spalte und Start-Zeile</li>
                    <li>Die Mini-Vorschau zeigt, wo die Daten eingefügt werden</li>
                    <li><b>Auto-Anfügen:</b> Nach letzter vorhandener Zeile einfügen</li>
                </ul>
            </li>
        </ul>

        <hr>
        <h3> Tipps</h3>
        <ul>
            <li>Spalten können per Drag & Drop umsortiert werden</li>
            <li>Rechtsklick auf Spalte → "Umbenennen" ändert nur den Export-Namen</li>
            <li>Bei "separate Blätter" wird jede Klasse auf ein eigenes Excel-Blatt exportiert</li>
            <li>UTF-8 Encoding wird verwendet - Umlaute funktionieren korrekt</li>
        </ul>
        """
        dialog = HelpDialog(" Hilfe - Excel Export", help_text, self)
        dialog.exec_()


class HelpDialog(QtWidgets.QDialog):
    def __init__(self, title: str, content_html: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(800, 600)
        
        self.setWindowFlags(
            QtCore.Qt.Window |
            QtCore.Qt.WindowMinimizeButtonHint |
            QtCore.Qt.WindowMaximizeButtonHint |
            QtCore.Qt.WindowCloseButtonHint
        )
        
        layout = QtWidgets.QVBoxLayout(self)
        
        text_edit = QtWidgets.QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setHtml(content_html)
        text_edit.setLineWrapMode(QtWidgets.QTextEdit.WidgetWidth)
        layout.addWidget(text_edit)
        
        button_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Close)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
class AdvancedExcelExportDialog(QtWidgets.QDialog):
    """
    Advanced Excel export dialog with full user control:
    - Select classes to export
    - Choose and reorder columns
    - Rename columns
    - Configure formatting
    - Save/load templates
    """
    
    def __init__(self, parent: 'WorkspaceWidget'):
        super().__init__(parent)
        self.workspace = parent
        self.setWindowTitle("Excel Export - Erweiterte Konfiguration")
        self.resize(900, 700)
        
        # ADD THESE LINES: Enable maximize button
        self.setWindowFlags(
            QtCore.Qt.Window |
            QtCore.Qt.WindowMinimizeButtonHint |
            QtCore.Qt.WindowMaximizeButtonHint |
            QtCore.Qt.WindowCloseButtonHint
        )
        
        # Template manager
        self.template_manager = ExportTemplateManager()
        
        self._build_ui()
        self._load_default_config()
    
    def _build_ui(self):
        main_layout = QtWidgets.QVBoxLayout(self)

        # --- HILFE-BUTTON HINZUFÜGEN (KORRIGIERTE PLATZIERUNG) ---
        # Erstellen Sie einen Container für die Überschrift und den Hilfe-Button
        header_container = QtWidgets.QWidget()
        header_layout = QtWidgets.QHBoxLayout(header_container)
        header_layout.setContentsMargins(0, 0, 0, 0) # Entfernt Standard-Ränder des Layouts

        help_button = QtWidgets.QPushButton("?")
        help_button.setFixedSize(24, 24)
        help_button.setToolTip("Hilfe zu diesem Dialog")
        help_button.clicked.connect(self._show_help)
        
        header_layout.addWidget(QtWidgets.QLabel("<h2>Excel Export Konfiguration</h2>"))
        header_layout.addStretch() # Sorgt dafür, dass der Button nach rechts geschoben wird
        header_layout.addWidget(help_button)
        
        main_layout.addWidget(header_container) # Fügen Sie den Container zum Hauptlayout hinzu
        # --- ENDE HILFE-BUTTON ---

        # ============================================================
        # SECTION 1: Template Management
        # ============================================================
        template_group = QtWidgets.QGroupBox(" Vorlagen")
        template_layout = QtWidgets.QHBoxLayout(template_group)
        
        template_layout.addWidget(QtWidgets.QLabel("Vorlage:"))
        
        self.template_combo = QtWidgets.QComboBox()
        self.template_combo.setMinimumWidth(200)
        self.template_combo.addItem("(Neue Konfiguration)")
        
        # Load saved templates
        for template_name in self.template_manager.list_templates():
            self.template_combo.addItem(template_name)
        
        self.template_combo.currentTextChanged.connect(self._on_template_changed)
        template_layout.addWidget(self.template_combo)
        
        self.btn_load_template = QtWidgets.QPushButton("Laden")
        self.btn_load_template.clicked.connect(self._load_template)
        template_layout.addWidget(self.btn_load_template)
        
        self.btn_save_template = QtWidgets.QPushButton("Speichern als...")
        self.btn_save_template.clicked.connect(self._save_template)
        template_layout.addWidget(self.btn_save_template)
        
        self.btn_delete_template = QtWidgets.QPushButton("Löschen")
        self.btn_delete_template.clicked.connect(self._delete_template)
        template_layout.addWidget(self.btn_delete_template)
        
        template_layout.addStretch()
        main_layout.addWidget(template_group) 
        
        # ============================================================
        # SECTION 2: Class Selection and Column Configuration
        # ============================================================
        config_splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        
        # Left: Class list
        left_widget = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_widget)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        left_layout.addWidget(QtWidgets.QLabel("Klassen auswählen:"))
        
        self.class_list = QtWidgets.QListWidget()
        self.class_list.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.class_list.currentRowChanged.connect(self._on_class_selected)
        self.class_list.itemChanged.connect(self._on_class_checkbox_changed)
        left_layout.addWidget(self.class_list)
        
        # Buttons for class selection
        class_btn_layout = QtWidgets.QHBoxLayout()
        self.btn_select_all_classes = QtWidgets.QPushButton("Alle")
        self.btn_select_all_classes.clicked.connect(self._select_all_classes)
        class_btn_layout.addWidget(self.btn_select_all_classes)
        
        self.btn_deselect_all_classes = QtWidgets.QPushButton("Keine")
        self.btn_deselect_all_classes.clicked.connect(self._deselect_all_classes)
        class_btn_layout.addWidget(self.btn_deselect_all_classes)
        
        left_layout.addLayout(class_btn_layout)
        
        config_splitter.addWidget(left_widget)
        
        # Right: Column configuration for selected class
        right_widget = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        self.current_class_label = QtWidgets.QLabel("Spalten für die ausgewählte Klasse konfigurieren:")
        self.current_class_label.setStyleSheet("font-weight: bold;")
        right_layout.addWidget(self.current_class_label)
        
        # Column configuration table
        self.column_table = QtWidgets.QTableWidget()
        self.column_table.setColumnCount(4)
        self.column_table.setHorizontalHeaderLabels([
            "Exportieren", "Spaltenname (Daten)", "Anzeigename (Excel)", "Reihenfolge"
        ])
        self.column_table.horizontalHeader().setStretchLastSection(False)
        self.column_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        self.column_table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        self.column_table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        self.column_table.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        self.column_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        right_layout.addWidget(self.column_table)
        
        # Column control buttons
        col_btn_layout = QtWidgets.QHBoxLayout()
        
        self.btn_select_all_cols = QtWidgets.QPushButton("Alle auswählen")
        self.btn_select_all_cols.clicked.connect(self._select_all_columns)
        col_btn_layout.addWidget(self.btn_select_all_cols)
        
        self.btn_deselect_all_cols = QtWidgets.QPushButton("Alle abwählen")
        self.btn_deselect_all_cols.clicked.connect(self._deselect_all_columns)
        col_btn_layout.addWidget(self.btn_deselect_all_cols)
        
        col_btn_layout.addStretch()
        
        self.btn_move_up = QtWidgets.QPushButton("↑ Nach oben")
        self.btn_move_up.clicked.connect(self._move_column_up)
        col_btn_layout.addWidget(self.btn_move_up)
        
        self.btn_move_down = QtWidgets.QPushButton("↓ Nach unten")
        self.btn_move_down.clicked.connect(self._move_column_down)
        col_btn_layout.addWidget(self.btn_move_down)
        
        right_layout.addLayout(col_btn_layout)
        
        config_splitter.addWidget(right_widget)
        config_splitter.setSizes([250, 650])
        
        main_layout.addWidget(config_splitter) 
        
        # ============================================================
        # SECTION 3: Export Options
        # ============================================================
        options_group = QtWidgets.QGroupBox(" Export-Optionen")
        options_layout = QtWidgets.QGridLayout(options_group)
        
        # Row 0: Sheet organization
        self.radio_separate_sheets = QtWidgets.QRadioButton("Separate Arbeitsblätter pro Klasse")
        self.radio_separate_sheets.setChecked(True)
        options_layout.addWidget(self.radio_separate_sheets, 0, 0, 1, 2)
        
        self.radio_single_sheet = QtWidgets.QRadioButton("Alle Klassen in einem Arbeitsblatt")
        options_layout.addWidget(self.radio_single_sheet, 1, 0, 1, 2)
        
        # Row 2: Filters
        self.check_apply_filters = QtWidgets.QCheckBox("Aktuelle Filter anwenden (Konfidenz, Sichtbarkeit)")
        self.check_apply_filters.setChecked(True)
        options_layout.addWidget(self.check_apply_filters, 2, 0, 1, 2)
        
        # Row 3: Empty cells
        self.check_include_empty = QtWidgets.QCheckBox("Leere Zellen einschließen")
        self.check_include_empty.setChecked(False)
        options_layout.addWidget(self.check_include_empty, 3, 0, 1, 2)
        
        # Row 4: Formatting
        self.check_auto_width = QtWidgets.QCheckBox("Spaltenbreite automatisch anpassen")
        self.check_auto_width.setChecked(True)
        options_layout.addWidget(self.check_auto_width, 4, 0)
        
        self.check_freeze_header = QtWidgets.QCheckBox("Kopfzeile fixieren")
        self.check_freeze_header.setChecked(True)
        options_layout.addWidget(self.check_freeze_header, 4, 1)
        
        # Row 5: Sorting
        options_layout.addWidget(QtWidgets.QLabel("Sortierung:"), 5, 0)
        self.sort_combo = QtWidgets.QComboBox()
        self.sort_combo.addItems([
            "Keine Sortierung",
            "Nach Klasse",
            "Nach Seite",
            "Nach Koordinatenwert",
            "Nach Text/Nummer"
        ])
        options_layout.addWidget(self.sort_combo, 5, 1)
        
        main_layout.addWidget(options_group)

        # ============================================================
        # SECTION 3.5: Export Target (New/Existing File)
        # ============================================================
        target_group = QtWidgets.QGroupBox(" Export-Ziel")
        target_layout = QtWidgets.QVBoxLayout(target_group)

        # Radio buttons for new vs existing file
        self.radio_new_file = QtWidgets.QRadioButton("Neue Datei erstellen")
        self.radio_new_file.setChecked(True)
        self.radio_new_file.toggled.connect(self._on_target_mode_changed)
        target_layout.addWidget(self.radio_new_file)

        self.radio_existing_file = QtWidgets.QRadioButton("In bestehende Datei einfügen")
        self.radio_existing_file.toggled.connect(self._on_target_mode_changed)
        target_layout.addWidget(self.radio_existing_file)

        # Container for existing file options (initially hidden)
        self.existing_file_container = QtWidgets.QWidget()
        existing_layout = QtWidgets.QGridLayout(self.existing_file_container)
        existing_layout.setContentsMargins(20, 5, 0, 5)

        # Row 0: File selection
        existing_layout.addWidget(QtWidgets.QLabel("Datei:"), 0, 0)
        self.existing_file_path = QtWidgets.QLineEdit()
        self.existing_file_path.setPlaceholderText("Excel-Datei auswählen...")
        self.existing_file_path.setReadOnly(True)
        existing_layout.addWidget(self.existing_file_path, 0, 1)

        self.btn_browse_file = QtWidgets.QPushButton("Durchsuchen...")
        self.btn_browse_file.clicked.connect(self._browse_existing_file)
        existing_layout.addWidget(self.btn_browse_file, 0, 2)

        # Row 1: Sheet selection
        existing_layout.addWidget(QtWidgets.QLabel("Arbeitsblatt:"), 1, 0)
        self.sheet_combo = QtWidgets.QComboBox()
        self.sheet_combo.setMinimumWidth(200)
        self.sheet_combo.setEnabled(False)
        existing_layout.addWidget(self.sheet_combo, 1, 1, 1, 2)

        # Row 2: Start position
        existing_layout.addWidget(QtWidgets.QLabel("Startposition:"), 2, 0)

        position_widget = QtWidgets.QWidget()
        position_layout = QtWidgets.QHBoxLayout(position_widget)
        position_layout.setContentsMargins(0, 0, 0, 0)

        position_layout.addWidget(QtWidgets.QLabel("Spalte:"))
        self.start_column = QtWidgets.QLineEdit("A")
        self.start_column.setMaximumWidth(50)
        self.start_column.setPlaceholderText("A")
        position_layout.addWidget(self.start_column)

        position_layout.addWidget(QtWidgets.QLabel("Zeile:"))
        self.start_row = QtWidgets.QSpinBox()
        self.start_row.setMinimum(1)
        self.start_row.setMaximum(1048576)  # Excel max rows
        self.start_row.setValue(1)
        position_layout.addWidget(self.start_row)

        position_layout.addStretch()
        existing_layout.addWidget(position_widget, 2, 1, 1, 2)

        # Row 3: Smart append option
        self.check_smart_append = QtWidgets.QCheckBox("Automatisch anfügen (nach letzter Zeile mit Daten)")
        self.check_smart_append.setToolTip(
            "Wenn aktiviert, werden die Daten automatisch nach der letzten "
            "vorhandenen Zeile eingefügt. Die Startzeile wird ignoriert."
        )
        self.check_smart_append.toggled.connect(self._on_smart_append_toggled)
        existing_layout.addWidget(self.check_smart_append, 3, 0, 1, 3)

        # Row 4: Header options for existing file
        self.check_include_header_existing = QtWidgets.QCheckBox("Kopfzeile mit exportieren")
        self.check_include_header_existing.setChecked(False)
        self.check_include_header_existing.setToolTip(
            "Wenn aktiviert, wird eine Kopfzeile mit Spaltennamen eingefügt. "
            "Deaktivieren Sie dies, wenn die Vorlage bereits Überschriften hat."
        )
        existing_layout.addWidget(self.check_include_header_existing, 4, 0, 1, 3)

        # Row 5: Preserve formatting option
        self.check_preserve_formatting = QtWidgets.QCheckBox("Formatierung der Vorlage beibehalten")
        self.check_preserve_formatting.setChecked(True)
        self.check_preserve_formatting.setToolTip(
            "Behält Zellformatierungen, Formeln und Stile der bestehenden Datei bei."
        )
        existing_layout.addWidget(self.check_preserve_formatting, 5, 0, 1, 3)

        target_layout.addWidget(self.existing_file_container)
        self.existing_file_container.setVisible(False)

        main_layout.addWidget(target_group)

        # ============================================================
        # SECTION 4: Data Preview Table
        # ============================================================
        preview_group = QtWidgets.QGroupBox(" Datenvorschau")
        preview_layout = QtWidgets.QVBoxLayout(preview_group)

        # Status label
        self.preview_status = QtWidgets.QLabel("Wählen Sie Klassen und Spalten aus, um eine Vorschau zu sehen.")
        self.preview_status.setStyleSheet("color: #666; font-style: italic;")
        preview_layout.addWidget(self.preview_status)

        # Data preview table
        self.preview_table = QtWidgets.QTableWidget()
        self.preview_table.setMinimumHeight(150)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.preview_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.preview_table.horizontalHeader().setStretchLastSection(True)
        preview_layout.addWidget(self.preview_table)

        # Preview controls
        preview_controls = QtWidgets.QHBoxLayout()

        self.btn_update_preview = QtWidgets.QPushButton(" Vorschau aktualisieren")
        self.btn_update_preview.clicked.connect(self._update_data_preview)
        preview_controls.addWidget(self.btn_update_preview)

        preview_controls.addStretch()

        self.preview_row_count = QtWidgets.QLabel("")
        self.preview_row_count.setStyleSheet("font-weight: bold;")
        preview_controls.addWidget(self.preview_row_count)

        preview_layout.addLayout(preview_controls)

        main_layout.addWidget(preview_group)

        # ============================================================
        # SECTION 5: Action Buttons
        # ============================================================
        button_layout = QtWidgets.QHBoxLayout()

        self.btn_export = QtWidgets.QPushButton(" Exportieren")
        self.btn_export.setStyleSheet("font-weight: bold; padding: 8px;")
        self.btn_export.clicked.connect(self._on_export_clicked)
        button_layout.addWidget(self.btn_export)

        self.btn_cancel = QtWidgets.QPushButton("Abbrechen")
        self.btn_cancel.clicked.connect(self.reject)
        button_layout.addWidget(self.btn_cancel)

        main_layout.addLayout(button_layout)

        # Initialize class list
        self._populate_class_list()

        # Initial preview update
        QtCore.QTimer.singleShot(100, self._update_data_preview)
    
    def _populate_class_list(self):
        """Populate the class list with checkboxes"""
        self.class_list.clear()
        self.class_items = {}
        
        if self.workspace.df_all is None or self.workspace.df_all.empty:
            return

        if 'cls' not in self.workspace.df_all.columns:
            return
        classes = sorted(self.workspace.df_all['cls'].unique())
        
        # Move 'coordinate' to end
        if 'coordinate' in classes:
            classes.remove('coordinate')
            classes.append('coordinate')
        
        for cls_name in classes:
            item = QtWidgets.QListWidgetItem(cls_name)
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Checked)  # Default: all checked
            self.class_list.addItem(item)
            self.class_items[cls_name] = item
        
        # Select first class
        if self.class_list.count() > 0:
            self.class_list.setCurrentRow(0)
    
    def _on_class_selected(self, row: int):
        """When a class is selected, show its column configuration"""
        if row < 0:
            return
        
        item = self.class_list.item(row)
        cls_name = item.text()
        
        self.current_class_label.setText(f"Spalten für '{cls_name}' konfigurieren:")
        self._populate_column_table(cls_name)
    
    def _populate_column_table(self, cls_name: str):
        """Populate column configuration table for a class"""
        self.column_table.setRowCount(0)
        
        df_class = self.workspace.df_all[self.workspace.df_all['cls'] == cls_name]
        if df_class.empty:
            return
        
        # Get available columns
        available_columns = self._get_available_columns(cls_name, df_class)
        
        # Default selected columns
        default_selected = self._get_default_columns(cls_name)
        
        # Populate table
        for idx, (col_key, col_display) in enumerate(available_columns.items()):
            row = self.column_table.rowCount()
            self.column_table.insertRow(row)
            
            # Column 0: Checkbox
            check_widget = QtWidgets.QWidget()
            check_layout = QtWidgets.QHBoxLayout(check_widget)
            check_layout.setContentsMargins(0, 0, 0, 0)
            check_layout.setAlignment(QtCore.Qt.AlignCenter)
            
            check = QtWidgets.QCheckBox()
            check.setChecked(col_key in default_selected)
            check_layout.addWidget(check)
            
            self.column_table.setCellWidget(row, 0, check_widget)
            
            # Column 1: Data column name (read-only)
            item_key = QtWidgets.QTableWidgetItem(col_key)
            item_key.setFlags(item_key.flags() & ~QtCore.Qt.ItemIsEditable)
            item_key.setData(QtCore.Qt.UserRole, col_key)  # Store original key
            self.column_table.setItem(row, 1, item_key)
            
            # Column 2: Display name (editable)
            item_display = QtWidgets.QTableWidgetItem(col_display)
            self.column_table.setItem(row, 2, item_display)
            
            # Column 3: Order (read-only, auto-updated)
            item_order = QtWidgets.QTableWidgetItem(str(idx + 1))
            item_order.setFlags(item_order.flags() & ~QtCore.Qt.ItemIsEditable)
            item_order.setTextAlignment(QtCore.Qt.AlignCenter)
            self.column_table.setItem(row, 3, item_order)
    
    def _get_available_columns(self, cls_name: str, df_class: pd.DataFrame) -> dict:
        """Get available columns for a class with display names"""
        if cls_name == 'coordinate':
            all_columns = {
                'coord_text': 'Koordinatentext',
                'coord_value': 'Koordinatenwert',
                'gi_gl': 'GI/GL',
                'page': 'Seite',
                'conf': 'Konfidenz',
                'color': 'Farbe',
                'notes': 'Notizen',
                'cx1': 'X1', 'cy1': 'Y1', 'cx2': 'X2', 'cy2': 'Y2',
                'angle': 'Winkel (normalisiert)',
                'angle_raw': 'Winkel (roh)',
                'obb_cx': 'OBB Center X',
                'obb_cy': 'OBB Center Y',
                'obb_w': 'OBB Breite',
                'obb_h': 'OBB Höhe',
            }
        else:
            all_columns = {
                'anchor_text': 'Text/Nummer',
                'coord_text': 'Koordinatentext',
                'coord_value': 'Koordinatenwert',
                'gi_gl': 'GI/GL',
                'page': 'Seite',
                'conf': 'Konfidenz',
                'color': 'Farbe',
                'notes': 'Notizen',
                'ax1': 'Anker X1', 'ay1': 'Anker Y1',
                'ax2': 'Anker X2', 'ay2': 'Anker Y2',
                'cx1': 'Koord X1', 'cy1': 'Koord Y1',
                'cx2': 'Koord X2', 'cy2': 'Koord Y2',
                'angle': 'Winkel (normalisiert)',
                'angle_raw': 'Winkel (roh)',
                'obb_cx': 'OBB Center X',
                'obb_cy': 'OBB Center Y',
                'obb_w': 'OBB Breite',
                'obb_h': 'OBB Höhe',
            }
            
            if cls_name == 'signal':
                all_columns['fahrtrichtung'] = 'Fahrtrichtung'
            
            if cls_name == 'weichen_block':
                all_columns['weichen_coordinates'] = 'Weichen Koordinaten'
        
        # Filter to only existing columns
        return {k: v for k, v in all_columns.items() if k in df_class.columns}
    
    def _get_default_columns(self, cls_name: str) -> list:
        """Get default selected columns for a class"""
        defaults = {
            'coordinate': ['coord_text', 'coord_value', 'gi_gl', 'page'],
            'signal': ['anchor_text', 'coord_text', 'coord_value', 'fahrtrichtung', 'page'],
            'gks_gesteuert': ['anchor_text', 'coord_text', 'coord_value', 'page'],
            'gks_festkodiert': ['anchor_text', 'coord_text', 'coord_value', 'page'],
            'weichen_block': ['anchor_text', 'weichen_coordinates', 'page'],
        }
        
        return defaults.get(cls_name, ['anchor_text', 'coord_text', 'page'])
    
    def _on_class_checkbox_changed(self, item):
        """Called when a class checkbox is toggled - update preview"""
        # Use a short delay to batch multiple changes
        if not hasattr(self, '_preview_timer'):
            self._preview_timer = QtCore.QTimer()
            self._preview_timer.setSingleShot(True)
            self._preview_timer.timeout.connect(self._update_data_preview)
        self._preview_timer.start(200)  # 200ms delay

    def _select_all_classes(self):
        """Select all classes"""
        for i in range(self.class_list.count()):
            item = self.class_list.item(i)
            item.setCheckState(QtCore.Qt.Checked)
        self._update_data_preview()

    def _deselect_all_classes(self):
        """Deselect all classes"""
        for i in range(self.class_list.count()):
            item = self.class_list.item(i)
            item.setCheckState(QtCore.Qt.Unchecked)
        self._update_data_preview()
    
    def _select_all_columns(self):
        """Select all columns in current table"""
        for row in range(self.column_table.rowCount()):
            widget = self.column_table.cellWidget(row, 0)
            if widget:
                check = widget.findChild(QtWidgets.QCheckBox)
                if check:
                    check.setChecked(True)
    
    def _deselect_all_columns(self):
        """Deselect all columns in current table"""
        for row in range(self.column_table.rowCount()):
            widget = self.column_table.cellWidget(row, 0)
            if widget:
                check = widget.findChild(QtWidgets.QCheckBox)
                if check:
                    check.setChecked(False)
    
    def _move_column_up(self):
        """Move selected column up"""
        current_row = self.column_table.currentRow()
        if current_row <= 0:
            return
        
        self._swap_rows(current_row, current_row - 1)
        self.column_table.setCurrentCell(current_row - 1, 0)
        self._update_order_numbers()
    
    def _move_column_down(self):
        """Move selected column down"""
        current_row = self.column_table.currentRow()
        if current_row < 0 or current_row >= self.column_table.rowCount() - 1:
            return
        
        self._swap_rows(current_row, current_row + 1)
        self.column_table.setCurrentCell(current_row + 1, 0)
        self._update_order_numbers()
    
    def _swap_rows(self, row1: int, row2: int):
        """Swap two rows in the column table"""
        for col in range(self.column_table.columnCount()):
            if col == 0:  # Checkbox column
                widget1 = self.column_table.cellWidget(row1, col)
                widget2 = self.column_table.cellWidget(row2, col)
                
                check1 = widget1.findChild(QtWidgets.QCheckBox)
                check2 = widget2.findChild(QtWidgets.QCheckBox)
                
                state1 = check1.isChecked()
                state2 = check2.isChecked()
                
                check1.setChecked(state2)
                check2.setChecked(state1)
            else:
                item1 = self.column_table.item(row1, col)
                item2 = self.column_table.item(row2, col)
                
                if item1 and item2:
                    text1 = item1.text()
                    text2 = item2.text()
                    data1 = item1.data(QtCore.Qt.UserRole)
                    data2 = item2.data(QtCore.Qt.UserRole)
                    
                    item1.setText(text2)
                    item2.setText(text1)
                    item1.setData(QtCore.Qt.UserRole, data2)
                    item2.setData(QtCore.Qt.UserRole, data1)
    
    def _update_order_numbers(self):
        """Update order numbers in column 3"""
        for row in range(self.column_table.rowCount()):
            item = self.column_table.item(row, 3)
            if item:
                item.setText(str(row + 1))

    def _update_data_preview(self):
        """Update the data preview table with actual data"""
        try:
            config = self.get_export_config()

            if not config['classes']:
                self.preview_table.setRowCount(0)
                self.preview_table.setColumnCount(0)
                self.preview_status.setText(" Keine Klassen ausgewählt!")
                self.preview_status.setStyleSheet("color: orange; font-weight: bold;")
                self.preview_row_count.setText("")
                return

            # Get base DataFrame
            df_base = self.workspace.df_all.copy()

            if df_base.empty:
                self.preview_status.setText(" Keine Daten vorhanden!")
                self.preview_status.setStyleSheet("color: red; font-weight: bold;")
                self.preview_row_count.setText("")
                return

            # Combine all selected classes
            all_rows = []
            all_columns = []
            display_names = {}

            for cls_name, cls_config in config['classes'].items():
                df_class = df_base[df_base['cls'] == cls_name].copy()

                if df_class.empty:
                    continue

                # Get selected columns
                cols = cls_config['columns']
                names = cls_config['display_names']

                # Filter to existing columns
                existing_cols = [c for c in cols if c in df_class.columns]

                if not existing_cols:
                    continue

                # Update column tracking
                for col in existing_cols:
                    if col not in all_columns:
                        all_columns.append(col)
                        display_names[col] = names.get(col, col)

                # Add rows
                for _, row in df_class.iterrows():
                    row_data = {'_class': cls_name}
                    for col in existing_cols:
                        val = row.get(col, '')
                        # Handle arrays/lists
                        if hasattr(val, '__iter__') and not isinstance(val, str):
                            val = str(list(val)) if len(val) > 0 else ''
                        elif pd.isna(val):
                            val = ''
                        row_data[col] = val
                    all_rows.append(row_data)

            if not all_rows:
                self.preview_status.setText(" Keine Daten nach Filterung!")
                self.preview_status.setStyleSheet("color: orange; font-weight: bold;")
                self.preview_row_count.setText("")
                self.preview_table.setRowCount(0)
                self.preview_table.setColumnCount(0)
                return

            # Limit preview to first 100 rows for performance
            max_preview_rows = 100
            preview_rows = all_rows[:max_preview_rows]
            total_rows = len(all_rows)

            # Add "Klasse" column at the beginning
            headers = ['Klasse'] + [display_names.get(c, c) for c in all_columns]

            # Setup table
            self.preview_table.setColumnCount(len(headers))
            self.preview_table.setRowCount(len(preview_rows))
            self.preview_table.setHorizontalHeaderLabels(headers)

            # Populate table
            for row_idx, row_data in enumerate(preview_rows):
                # Class column
                cls_item = QtWidgets.QTableWidgetItem(row_data.get('_class', ''))
                cls_item.setBackground(QtGui.QColor(240, 240, 240))
                self.preview_table.setItem(row_idx, 0, cls_item)

                # Data columns
                for col_idx, col in enumerate(all_columns):
                    val = row_data.get(col, '')
                    item = QtWidgets.QTableWidgetItem(str(val))
                    self.preview_table.setItem(row_idx, col_idx + 1, item)

            # Resize columns
            self.preview_table.resizeColumnsToContents()

            # Update status
            self.preview_status.setText(" Vorschau der Exportdaten:")
            self.preview_status.setStyleSheet("color: green; font-weight: bold;")

            if total_rows > max_preview_rows:
                self.preview_row_count.setText(
                    f" {total_rows} Zeilen gesamt (zeige erste {max_preview_rows})"
                )
            else:
                self.preview_row_count.setText(f" {total_rows} Zeilen werden exportiert")

        except Exception as e:
            self.preview_status.setText(f" Fehler: {str(e)}")
            self.preview_status.setStyleSheet("color: red;")
            self.preview_row_count.setText("")
            print(f"Preview error: {e}")

    def _load_default_config(self):
        """Load default configuration"""
        # Already done in _populate_class_list and _populate_column_table
        pass
    
    def _on_template_changed(self, template_name: str):
        """Handle template selection change"""
        if template_name == "(Neue Konfiguration)":
            self.btn_delete_template.setEnabled(False)
        else:
            self.btn_delete_template.setEnabled(True)
    
    def _load_template(self):
        """Load selected template"""
        template_name = self.template_combo.currentText()
        if template_name == "(Neue Konfiguration)":
            QtWidgets.QMessageBox.information(
                self,
                "Keine Vorlage",
                "Bitte wählen Sie eine gespeicherte Vorlage aus."
            )
            return
        
        config = self.template_manager.get_template(template_name)
        if not config:
            QtWidgets.QMessageBox.warning(
                self,
                "Fehler",
                f"Vorlage '{template_name}' konnte nicht geladen werden."
            )
            return
        
        self._apply_config(config)
        QtWidgets.QMessageBox.information(
            self,
            "Geladen",
            f"Vorlage '{template_name}' wurde geladen."
        )
    
    def _save_template(self):
        """Save current configuration as template"""
        name, ok = QtWidgets.QInputDialog.getText(
            self,
            "Vorlage speichern",
            "Name der Vorlage:",
            text=self.template_combo.currentText() if self.template_combo.currentText() != "(Neue Konfiguration)" else ""
        )
        
        if not ok or not name.strip():
            return
        
        name = name.strip()
        
        # Check if overwriting
        if name in self.template_manager.list_templates():
            reply = QtWidgets.QMessageBox.question(
                self,
                "Überschreiben?",
                f"Vorlage '{name}' existiert bereits. Überschreiben?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
            )
            
            if reply != QtWidgets.QMessageBox.Yes:
                return
        
        config = self.get_export_config()
        self.template_manager.save_template(name, config)
        
        # Add to combo if new
        if self.template_combo.findText(name) == -1:
            self.template_combo.addItem(name)
        
        self.template_combo.setCurrentText(name)
        
        QtWidgets.QMessageBox.information(
            self,
            "Gespeichert",
            f"Vorlage '{name}' wurde gespeichert."
        )
    
    def _delete_template(self):
        """Delete selected template"""
        template_name = self.template_combo.currentText()
        if template_name == "(Neue Konfiguration)":
            return
        
        reply = QtWidgets.QMessageBox.question(
            self,
            "Löschen?",
            f"Vorlage '{template_name}' wirklich löschen?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )
        
        if reply != QtWidgets.QMessageBox.Yes:
            return
        
        self.template_manager.delete_template(template_name)
        
        # Remove from combo
        idx = self.template_combo.findText(template_name)
        if idx >= 0:
            self.template_combo.removeItem(idx)
        
        self.template_combo.setCurrentIndex(0)
        
        QtWidgets.QMessageBox.information(
            self,
            "Gelöscht",
            f"Vorlage '{template_name}' wurde gelöscht."
        )
    
    def _apply_config(self, config: dict):
        """Apply a configuration to the UI"""
        # Apply class selections
        for cls_name, item in self.class_items.items():
            if cls_name in config.get('classes', {}):
                item.setCheckState(QtCore.Qt.Checked)
            else:
                item.setCheckState(QtCore.Qt.Unchecked)
        
        # Apply options
        opts = config.get('options', {})
        
        if opts.get('separate_sheets', True):
            self.radio_separate_sheets.setChecked(True)
        else:
            self.radio_single_sheet.setChecked(True)
        
        self.check_apply_filters.setChecked(opts.get('apply_filters', True))
        self.check_include_empty.setChecked(opts.get('include_empty', False))
        self.check_auto_width.setChecked(opts.get('auto_width', True))
        self.check_freeze_header.setChecked(opts.get('freeze_header', True))
        
        sort_text = opts.get('sort_by', 'Keine Sortierung')
        idx = self.sort_combo.findText(sort_text)
        if idx >= 0:
            self.sort_combo.setCurrentIndex(idx)
        
        # Refresh current class column config
        current_row = self.class_list.currentRow()
        if current_row >= 0:
            self._on_class_selected(current_row)
    
    def get_export_config(self) -> dict:
        """
        Get the complete export configuration.

        Returns dict with structure:
        {
            'classes': {
                'class_name': {
                    'columns': ['col1', 'col2', ...],
                    'display_names': {'col1': 'Display 1', ...},
                    'order': [0, 1, 2, ...]
                }
            },
            'options': {
                'separate_sheets': bool,
                'apply_filters': bool,
                'include_empty': bool,
                'auto_width': bool,
                'freeze_header': bool,
                'sort_by': str
            },
            'target': {
                'mode': 'new_file' | 'existing_file',
                ...  # additional fields for existing_file mode
            }
        }
        """
        config = {
            'classes': {},
            'options': {
                'separate_sheets': self.radio_separate_sheets.isChecked(),
                'apply_filters': self.check_apply_filters.isChecked(),
                'include_empty': self.check_include_empty.isChecked(),
                'auto_width': self.check_auto_width.isChecked(),
                'freeze_header': self.check_freeze_header.isChecked(),
                'sort_by': self.sort_combo.currentText(),
            },
            'target': self.get_target_config()
        }
        
        # Get configuration for each selected class
        for i in range(self.class_list.count()):
            item = self.class_list.item(i)
            if item.checkState() != QtCore.Qt.Checked:
                continue
            
            cls_name = item.text()
            
            # Temporarily switch to this class to get its config
            # (This is important to ensure column_table is populated correctly for the class)
            current_class_list_row = self.class_list.currentRow()
            self.class_list.setCurrentRow(i) 
            
            columns = []
            display_names = {}
            # order is implicitly handled by the order of 'columns' list
            
            for row in range(self.column_table.rowCount()):
                # Check if column is selected
                widget = self.column_table.cellWidget(row, 0)
                if not widget:
                    continue
                
                check = widget.findChild(QtWidgets.QCheckBox)
                if not check or not check.isChecked():
                    continue
                
                # Get column key
                item_key = self.column_table.item(row, 1)
                if not item_key:
                    continue
                
                col_key = item_key.data(QtCore.Qt.UserRole)
                if not col_key:
                    col_key = item_key.text()
                
                # Get display name
                item_display = self.column_table.item(row, 2)
                display_name = item_display.text() if item_display else col_key
                
                columns.append(col_key)
                display_names[col_key] = display_name
            
            if columns:
                config['classes'][cls_name] = {
                    'columns': columns,
                    'display_names': display_names,
                }
            
            # Restore original class selection
            self.class_list.setCurrentRow(current_class_list_row)
        
        return config
    
    def _show_help(self):
        """Show help for the AdvancedExcelExportDialog"""
        help_text = """
        <h2>Excel Export Konfiguration - Hilfe</h2>

        <h3>Zweck</h3>
        <p>Dieser Dialog ermöglicht Ihnen die detaillierte Konfiguration des Exports Ihrer extrahierten Daten nach Excel.</p>

        <h3>1. Vorlagen</h3>
        <ul>
            <li><b>Vorlage wählen:</b> Wählen Sie eine gespeicherte Vorlage aus der Dropdown-Liste.</li>
            <li><b>Laden:</b> Lädt die ausgewählte Vorlage und wendet ihre Einstellungen an.</li>
            <li><b>Speichern als...:</b> Speichert die aktuellen Einstellungen als neue Vorlage.</li>
            <li><b>Löschen:</b> Löscht die ausgewählte Vorlage.</li>
        </ul>

        <h3>2. Klassen & Spalten konfigurieren</h3>
        <h4>Linke Seite: Klassen auswählen</h4>
        <ul>
            <li>Wählen Sie die Klassen (z.B. Signal, Koordinate), die Sie in den Export aufnehmen möchten, indem Sie die Kontrollkästchen aktivieren/deaktivieren.</li>
            <li><b>Alle / Keine:</b> Buttons zum schnellen Auswählen oder Abwählen aller Klassen.</li>
        </ul>
        <h4>Rechte Seite: Spalten für die ausgewählte Klasse konfigurieren</h4>
        <ul>
            <li>Wählen Sie eine Klasse auf der linken Seite aus, um ihre Spalten hier zu konfigurieren.</li>
            <li><b>Exportieren:</b> Aktivieren/Deaktivieren Sie das Kontrollkästchen, um die Spalte in den Export aufzunehmen.</li>
            <li><b>Spaltenname (Daten):</b> Der interne Name des Datenfeldes (nicht bearbeitbar).</li>
            <li><b>Anzeigename (Excel):</b> Der Text, der als Spaltenüberschrift in Excel verwendet wird (bearbeitbar).</li>
            <li><b>Reihenfolge:</b> Zeigt die aktuelle Position der Spalte an. Verwenden Sie die Pfeil-Buttons (<b>↑ Nach oben / ↓ Nach unten</b>), um die Reihenfolge der Spalten im Export anzupassen.</li>
            <li><b>Alle auswählen / Alle abwählen:</b> Buttons zum schnellen Auswählen oder Abwählen aller Spalten für die aktuelle Klasse.</li>
        </ul>

        <h3>3. Export-Optionen</h3>
        <ul>
            <li><b>Separate Arbeitsblätter pro Klasse:</b>
                <ul>
                    <li><b>Aktiviert:</b> Jede ausgewählte Klasse wird in ein eigenes Arbeitsblatt in der Excel-Datei exportiert.</li>
                    <li><b>Deaktiviert:</b> Alle ausgewählten Daten werden in einem einzigen Arbeitsblatt zusammengeführt.</li>
                </ul>
            </li>
            <li><b>Aktuelle Filter anwenden:</b> Wenn aktiviert, werden nur die Elemente exportiert, die den aktuell im Auditing-Fenster eingestellten Filtern (Klasse, Text, Konfidenz) entsprechen.</li>
            <li><b>Leere Zellen einschließen:</b> Wenn deaktiviert, werden leere Werte als leere Zeichenketten exportiert.</li>
            <li><b>Spaltenbreite automatisch anpassen:</b> Versucht, die Spaltenbreite in Excel automatisch an den Inhalt anzupassen.</li>
            <li><b>Kopfzeile fixieren:</b> Friert die erste Zeile (Spaltenüberschriften) in Excel ein, sodass diese beim Scrollen sichtbar bleibt.</li>
            <li><b>Sortierung:</b> Legt die primäre Sortierreihenfolge der Daten im Excel-Export fest.</li>
        </ul>

        <h3>4. Export-Ziel (NEU)</h3>
        <p>Wählen Sie, ob Sie in eine <b>neue Datei</b> oder in eine <b>bestehende Excel-Datei</b> exportieren möchten.</p>

        <h4>Neue Datei erstellen</h4>
        <p>Erstellt eine neue Excel-Datei mit Ihren Daten. Sie werden nach dem Speicherort gefragt.</p>

        <h4>In bestehende Datei einfügen</h4>
        <p>Fügt die Daten in eine bereits vorhandene Excel-Datei ein. Ideal für Vorlagen oder fortlaufende Datenpflege.</p>
        <ul>
            <li><b>Datei:</b> Wählen Sie die bestehende Excel-Datei aus.</li>
            <li><b>Arbeitsblatt:</b> Wählen Sie das Ziel-Arbeitsblatt aus der Datei.</li>
            <li><b>Startposition:</b> Geben Sie die Spalte (z.B. A, B, AA) und Zeile (z.B. 1, 5, 100) an, ab der die Daten eingefügt werden sollen.</li>
            <li><b>Automatisch anfügen:</b> Wenn aktiviert, werden die Daten automatisch nach der letzten Zeile mit Inhalt eingefügt. Die Startzeile wird dann ignoriert.</li>
            <li><b>Kopfzeile mit exportieren:</b> Fügt eine Kopfzeile mit Spaltennamen ein. Deaktivieren Sie dies, wenn Ihre Vorlage bereits Überschriften hat.</li>
            <li><b>Formatierung beibehalten:</b> Behält vorhandene Formatierungen, Formeln und Stile der bestehenden Datei bei.</li>
        </ul>

        <h3>5. Vorschau</h3>
        <p>Zeigt eine Zusammenfassung Ihrer aktuellen Export-Einstellungen an. Klicken Sie auf <b>"Vorschau aktualisieren"</b>, um die Anzeige zu aktualisieren.</p>

        <h3>Export starten</h3>
        <p>Klicken Sie auf <b>" Exportieren"</b>, um den Export zu starten.</p>
        <ul>
            <li><b>Bei neuer Datei:</b> Sie werden nach dem Speicherort gefragt.</li>
            <li><b>Bei bestehender Datei:</b> Die ausgewählte Datei wird direkt aktualisiert (ein Backup wird empfohlen!).</li>
        </ul>
        """
        
        # Verwenden Sie den neuen HelpDialog
        help_dialog = HelpDialog("Hilfe - Excel Export Konfiguration", help_text, self)
        help_dialog.exec_()

    def _on_export_clicked(self):
        """Handle export button click with validation"""
        is_valid, error_message = self.validate_for_export()

        if not is_valid:
            QtWidgets.QMessageBox.warning(
                self,
                "Validierungsfehler",
                error_message
            )
            return

        self.accept()

    def _on_target_mode_changed(self, checked: bool):
        """Toggle visibility of existing file options"""
        if self.radio_existing_file.isChecked():
            self.existing_file_container.setVisible(True)
            # Disable some options that don't apply to existing files
            self.radio_separate_sheets.setEnabled(False)
            self.radio_single_sheet.setEnabled(False)
        else:
            self.existing_file_container.setVisible(False)
            self.radio_separate_sheets.setEnabled(True)
            self.radio_single_sheet.setEnabled(True)

    def _browse_existing_file(self):
        """Browse for an existing Excel file"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Bestehende Excel-Datei auswählen",
            "",
            "Excel Dateien (*.xlsx *.xls)"
        )

        if not file_path:
            return

        self.existing_file_path.setText(file_path)

        # Load sheet names from the file
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheet_names)
            self.sheet_combo.setEnabled(True)

            # If there's only one sheet, auto-select it
            if len(sheet_names) == 1:
                self.sheet_combo.setCurrentIndex(0)

        except Exception as e:
            QtWidgets.QMessageBox.warning(
                self,
                "Fehler beim Laden",
                f"Die Datei konnte nicht gelesen werden:\n{e}"
            )
            self.existing_file_path.clear()
            self.sheet_combo.clear()
            self.sheet_combo.setEnabled(False)

    def _on_smart_append_toggled(self, checked: bool):
        """Enable/disable start row when smart append is toggled"""
        self.start_row.setEnabled(not checked)
        if checked:
            self.start_row.setToolTip("Wird automatisch ermittelt")
        else:
            self.start_row.setToolTip("")

    def _validate_column_input(self, col_str: str) -> bool:
        """Validate that column input is valid Excel column (A-XFD)"""
        if not col_str:
            return False
        col_str = col_str.upper().strip()
        # Excel columns are A-XFD (1 to 16384)
        if not re.match(r'^[A-Z]{1,3}$', col_str):
            return False
        try:
            col_idx = column_index_from_string(col_str)
            return 1 <= col_idx <= 16384
        except:
            return False

    def get_target_config(self) -> dict:
        """Get export target configuration"""
        if self.radio_new_file.isChecked():
            return {
                'mode': 'new_file'
            }
        else:
            # Validate inputs
            col_str = self.start_column.text().upper().strip() or 'A'

            if not self._validate_column_input(col_str):
                col_str = 'A'

            return {
                'mode': 'existing_file',
                'file_path': self.existing_file_path.text(),
                'sheet_name': self.sheet_combo.currentText(),
                'start_column': col_str,
                'start_row': self.start_row.value(),
                'smart_append': self.check_smart_append.isChecked(),
                'include_header': self.check_include_header_existing.isChecked(),
                'preserve_formatting': self.check_preserve_formatting.isChecked()
            }

    def validate_for_export(self) -> tuple:
        """
        Validate configuration before export.
        Returns (is_valid: bool, error_message: str)
        """
        target = self.get_target_config()

        if target['mode'] == 'existing_file':
            if not target['file_path']:
                return False, "Bitte wählen Sie eine bestehende Excel-Datei aus."

            if not os.path.exists(target['file_path']):
                return False, f"Die Datei existiert nicht:\n{target['file_path']}"

            if not target['sheet_name']:
                return False, "Bitte wählen Sie ein Arbeitsblatt aus."

            if not self._validate_column_input(self.start_column.text()):
                return False, "Ungültige Spaltenangabe. Bitte geben Sie einen gültigen Excel-Spaltenbuchstaben ein (z.B. A, B, AA)."

        # Check that at least one class is selected
        config = self.get_export_config()
        if not config['classes']:
            return False, "Bitte wählen Sie mindestens eine Klasse mit Spalten aus."

        return True, ""


class ExportTemplateManager:
    """Manage export templates with JSON persistence"""
    
    def __init__(self):
        self.templates_file = "export_templates.json"
        self.templates = self._load_templates()
    
    def _load_templates(self) -> dict:
        """Load templates from file"""
        if os.path.exists(self.templates_file):
            try:
                with open(self.templates_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Failed to load templates: {e}")
                return {}
        return {}
    
    def _save_templates(self):
        """Save templates to file"""
        try:
            with open(self.templates_file, 'w', encoding='utf-8') as f:
                json.dump(self.templates, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Failed to save templates: {e}")
    
    def save_template(self, name: str, config: dict):
        """Save a template"""
        self.templates[name] = config
        self._save_templates()
    
    def get_template(self, name: str) -> dict:
        """Get a template by name"""
        return self.templates.get(name, {})
    
    def delete_template(self, name: str):
        """Delete a template"""
        if name in self.templates:
            del self.templates[name]
            self._save_templates()
    
    def list_templates(self) -> list:
        """List all template names"""
        return list(self.templates.keys())