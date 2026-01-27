"""
Shared Export Utilities

This module provides reusable export functions for DataFrames.
Used by:
- SimpleExcelExportDialog (main export)
- SimplePDFCompareDialog (comparison export)
- Any future export needs
"""

import os
import json
import pandas as pd
from typing import Dict, List, Optional, Any
from dataclasses import dataclass


@dataclass
class ExportOptions:
    """Options for export operations"""
    include_header: bool = True
    auto_width: bool = True
    freeze_header: bool = True
    apply_filters: bool = False
    empty_cell_value: str = ''
    number_format: str = 'Original'  # 'Original', '2 Dezimalen', 'Ganzzahl'
    csv_separator: str = ';'
    json_orient: str = 'records'  # 'records', 'grouped', 'table', 'index'
    json_indent: int = 2
    sheet_name: str = 'Export'


class ExportError(Exception):
    """Custom exception for export errors"""
    pass


def export_dataframe(
    df: pd.DataFrame,
    file_path: str,
    format_type: str = 'xlsx',
    options: Optional[ExportOptions] = None
) -> bool:
    """
    Export a DataFrame to file.

    Args:
        df: DataFrame to export
        file_path: Output file path
        format_type: 'xlsx', 'csv', or 'json'
        options: Export options

    Returns:
        True if successful

    Raises:
        ExportError: If export fails
    """
    if options is None:
        options = ExportOptions()

    if df.empty:
        raise ExportError("Keine Daten zum Exportieren")

    # Normalize format type
    fmt = format_type.lower()
    if 'xlsx' in fmt or 'excel' in fmt:
        return export_to_excel(df, file_path, options)
    elif 'csv' in fmt:
        return export_to_csv(df, file_path, options)
    elif 'json' in fmt:
        return export_to_json(df, file_path, options)
    else:
        raise ExportError(f"Unbekanntes Format: {format_type}")


def export_to_excel(
    df: pd.DataFrame,
    file_path: str,
    options: Optional[ExportOptions] = None
) -> bool:
    """
    Export DataFrame to Excel file.

    Args:
        df: DataFrame to export
        file_path: Output file path (.xlsx)
        options: Export options

    Returns:
        True if successful
    """
    if options is None:
        options = ExportOptions()

    # Ensure .xlsx extension
    if not file_path.endswith('.xlsx'):
        file_path += '.xlsx'

    # Apply formatting
    df = _apply_formatting(df, options)

    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = options.sheet_name[:31]  # Excel sheet name max 31 chars

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=options.include_header), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Auto-adjust column widths
        if options.auto_width:
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        if options.freeze_header and options.include_header:
            ws.freeze_panes = 'A2'

        # Apply auto-filter
        if options.apply_filters and options.include_header and len(df) > 0:
            last_col = get_column_letter(len(df.columns))
            last_row = len(df) + 1
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        wb.save(file_path)
        return True

    except PermissionError:
        raise ExportError(
            f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
            "Die Datei ist wahrscheinlich in Excel geöffnet.\n"
            "Bitte schließen Sie die Datei und versuchen Sie es erneut."
        )
    except Exception as e:
        raise ExportError(f"Excel Export fehlgeschlagen: {str(e)}")


def export_to_excel_multi_sheet(
    dataframes: Dict[str, pd.DataFrame],
    file_path: str,
    options: Optional[ExportOptions] = None
) -> bool:
    """
    Export multiple DataFrames to separate sheets in one Excel file.

    Args:
        dataframes: Dict of {sheet_name: DataFrame}
        file_path: Output file path (.xlsx)
        options: Export options

    Returns:
        True if successful
    """
    if options is None:
        options = ExportOptions()

    if not file_path.endswith('.xlsx'):
        file_path += '.xlsx'

    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        sheets_created = 0
        for sheet_name, df in dataframes.items():
            if df.empty:
                continue

            df = _apply_formatting(df, options)

            # Truncate sheet name to 31 chars
            safe_name = sheet_name[:31]
            ws = wb.create_sheet(title=safe_name)

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=options.include_header), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Auto-adjust column widths
            if options.auto_width:
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header
            if options.freeze_header and options.include_header:
                ws.freeze_panes = 'A2'

            sheets_created += 1

        if sheets_created == 0:
            raise ExportError("Keine Daten zum Exportieren")

        wb.save(file_path)
        return True

    except PermissionError:
        raise ExportError(
            f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
            "Die Datei ist wahrscheinlich in Excel geöffnet."
        )
    except ExportError:
        raise
    except Exception as e:
        raise ExportError(f"Excel Export fehlgeschlagen: {str(e)}")


def export_to_csv(
    df: pd.DataFrame,
    file_path: str,
    options: Optional[ExportOptions] = None
) -> bool:
    """
    Export DataFrame to CSV file.

    Args:
        df: DataFrame to export
        file_path: Output file path (.csv)
        options: Export options

    Returns:
        True if successful
    """
    if options is None:
        options = ExportOptions()

    if not file_path.endswith('.csv'):
        file_path += '.csv'

    df = _apply_formatting(df, options)

    try:
        df.to_csv(
            file_path,
            index=False,
            sep=options.csv_separator,
            encoding='utf-8-sig',  # UTF-8 with BOM for Excel compatibility
            header=options.include_header
        )
        return True

    except PermissionError:
        raise ExportError(
            f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
            "Die Datei ist möglicherweise geöffnet."
        )
    except Exception as e:
        raise ExportError(f"CSV Export fehlgeschlagen: {str(e)}")


def export_to_json(
    df: pd.DataFrame,
    file_path: str,
    options: Optional[ExportOptions] = None
) -> bool:
    """
    Export DataFrame to JSON file.

    Args:
        df: DataFrame to export
        file_path: Output file path (.json)
        options: Export options

    Returns:
        True if successful
    """
    if options is None:
        options = ExportOptions()

    if not file_path.endswith('.json'):
        file_path += '.json'

    df = _apply_formatting(df, options)

    try:
        orient = options.json_orient.lower()

        if orient == 'records':
            export_data = df.to_dict(orient='records')
        elif orient == 'table':
            export_data = {
                'columns': list(df.columns),
                'data': df.values.tolist()
            }
        elif orient == 'index':
            export_data = df.to_dict(orient='index')
        else:
            export_data = df.to_dict(orient='records')

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=options.json_indent, default=str)

        return True

    except PermissionError:
        raise ExportError(
            f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
            "Die Datei ist möglicherweise geöffnet."
        )
    except Exception as e:
        raise ExportError(f"JSON Export fehlgeschlagen: {str(e)}")


def export_to_json_grouped(
    dataframes: Dict[str, pd.DataFrame],
    file_path: str,
    options: Optional[ExportOptions] = None
) -> bool:
    """
    Export multiple DataFrames to a single JSON file, grouped by key.

    Args:
        dataframes: Dict of {group_name: DataFrame}
        file_path: Output file path (.json)
        options: Export options

    Returns:
        True if successful
    """
    if options is None:
        options = ExportOptions()

    if not file_path.endswith('.json'):
        file_path += '.json'

    try:
        export_data = {}
        for group_name, df in dataframes.items():
            if not df.empty:
                df = _apply_formatting(df, options)
                export_data[group_name] = df.to_dict(orient='records')

        if not export_data:
            raise ExportError("Keine Daten zum Exportieren")

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=options.json_indent, default=str)

        return True

    except PermissionError:
        raise ExportError(
            f"Die Datei kann nicht gespeichert werden:\n{file_path}\n\n"
            "Die Datei ist möglicherweise geöffnet."
        )
    except ExportError:
        raise
    except Exception as e:
        raise ExportError(f"JSON Export fehlgeschlagen: {str(e)}")


def _apply_formatting(df: pd.DataFrame, options: ExportOptions) -> pd.DataFrame:
    """Apply formatting options to DataFrame"""
    df = df.copy()

    # Handle empty cells
    if options.empty_cell_value:
        df = df.fillna(options.empty_cell_value)
        df = df.replace('', options.empty_cell_value)

    # Number formatting
    if options.number_format != 'Original':
        for col in df.columns:
            try:
                numeric_col = pd.to_numeric(df[col], errors='coerce')
                mask = numeric_col.notna()
                if mask.any():
                    if options.number_format == '2 Dezimalen':
                        df.loc[mask, col] = numeric_col[mask].apply(lambda x: f"{x:.2f}")
                    elif options.number_format == 'Ganzzahl':
                        df.loc[mask, col] = numeric_col[mask].apply(lambda x: str(int(x)))
            except:
                pass

    return df


# =============================================================================
# COMPARISON-SPECIFIC EXPORT HELPERS
# =============================================================================

def comparison_results_to_dataframes(results: Dict) -> Dict[str, pd.DataFrame]:
    """
    Convert comparison results to DataFrames for export.

    Args:
        results: Comparison results from LayoutComparisonEngine.compare()
                 Contains 'added', 'deleted', 'moved', 'modified', 'summary' keys

    Returns:
        Dict with DataFrames for each result type (FA-011):
        - 'Zusammenfassung': Summary statistics
        - 'Hinzugefügt': Added elements
        - 'Gelöscht': Deleted elements
        - 'Verschoben': Moved elements (same identifier, different position)
        - 'Modifiziert': Modified elements with field-level changes
    """
    dataframes = {}

    # Summary DataFrame
    if 'summary' in results:
        summary = results['summary']
        summary_data = [
            {'Kategorie': 'Hinzugefügt', 'Anzahl': summary.get('total_added', 0)},
            {'Kategorie': 'Gelöscht', 'Anzahl': summary.get('total_deleted', 0)},
            {'Kategorie': 'Verschoben', 'Anzahl': summary.get('total_moved', 0)},  # FA-011
            {'Kategorie': 'Modifiziert', 'Anzahl': summary.get('total_modified', 0)},
            {'Kategorie': 'Unverändert', 'Anzahl': summary.get('total_unchanged', 0)},
        ]

        # Add severity breakdown
        by_severity = summary.get('by_severity', {})
        if by_severity:
            summary_data.append({'Kategorie': '---', 'Anzahl': '---'})
            summary_data.append({'Kategorie': 'Nach Schweregrad:', 'Anzahl': ''})
            for sev, count in by_severity.items():
                summary_data.append({'Kategorie': f'  {sev.capitalize()}', 'Anzahl': count})

        dataframes['Zusammenfassung'] = pd.DataFrame(summary_data)

    # Added elements (FA-011: Hinzugefügt)
    if 'added' in results and results['added']:
        added_data = []
        for change in results['added']:
            data = change.new_data or {}
            added_data.append({
                'Objektklasse': data.get('cls', ''),
                'Kennung': data.get('anchor_text', '') or data.get('coord_text', ''),
                'Änderungstyp': 'Hinzugefügt',
                'Alt-Wert': '',
                'Neu-Wert': data.get('anchor_text', ''),
                'Koordinate': data.get('coord_text', ''),
                'Koordinatenwert': _format_coord_value(data.get('coord_value')),
                'Seite': data.get('page', ''),
            })
        dataframes['Hinzugefügt'] = pd.DataFrame(added_data)

    # Deleted elements (FA-011: Entfernt)
    if 'deleted' in results and results['deleted']:
        deleted_data = []
        for change in results['deleted']:
            data = change.old_data or {}
            deleted_data.append({
                'Objektklasse': data.get('cls', ''),
                'Kennung': data.get('anchor_text', '') or data.get('coord_text', ''),
                'Änderungstyp': 'Entfernt',
                'Alt-Wert': data.get('anchor_text', ''),
                'Neu-Wert': '',
                'Koordinate': data.get('coord_text', ''),
                'Koordinatenwert': _format_coord_value(data.get('coord_value')),
                'Seite': data.get('page', ''),
            })
        dataframes['Gelöscht'] = pd.DataFrame(deleted_data)

    # Moved elements (FA-011: Verschoben - Same identifier, different km position)
    if 'moved' in results and results['moved']:
        moved_data = []
        for change in results['moved']:
            old_data = change.old_data or {}
            new_data = change.new_data or {}
            spatial = change.spatial_change or {}

            # Format coord_value change (km position along track)
            old_coord = spatial.get('old_coord')
            new_coord = spatial.get('new_coord')
            delta_m = spatial.get('delta_meters')

            old_str = _format_coord_value(old_coord)
            new_str = _format_coord_value(new_coord)
            if delta_m is not None:
                sign = '+' if delta_m > 0 else ''
                new_str += f" ({sign}{delta_m:.1f}m)"

            moved_data.append({
                'Objektklasse': new_data.get('cls', ''),
                'Kennung': new_data.get('anchor_text', '') or new_data.get('coord_text', ''),
                'Änderungstyp': 'Verschoben',
                'Alt-Wert': old_str,
                'Neu-Wert': new_str,
                'Koordinate': new_data.get('coord_text', ''),
                'Seite': new_data.get('page', ''),
            })
        dataframes['Verschoben'] = pd.DataFrame(moved_data)

    # Modified elements (FA-011: Modifiziert - one row per changed field)
    if 'modified' in results and results['modified']:
        modified_data = []
        for change in results['modified']:
            if not change.field_changes:
                continue

            old_data = change.old_data or {}
            new_data = change.new_data or {}

            for field, diff in change.field_changes.items():
                old_val = diff.get('old', '')
                new_val = diff.get('new', '')

                # Format coordinate values with delta
                if field == 'coord_value':
                    old_str = _format_coord_value(old_val)
                    new_str = _format_coord_value(new_val)
                    if 'delta_meters' in diff and diff['delta_meters'] is not None:
                        delta_m = diff['delta_meters']
                        sign = '+' if delta_m > 0 else ''
                        new_str += f" ({sign}{delta_m:.1f}m)"
                else:
                    old_str = str(old_val) if old_val is not None else ''
                    new_str = str(new_val) if new_val is not None else ''

                # Human-readable field names for Änderungstyp
                field_names = {
                    'cls': 'Klasse geändert',
                    'anchor_text': 'Text geändert',
                    'coord_text': 'Koordinatentext geändert',
                    'coord_value': 'Position geändert',
                    'fahrtrichtung': 'Fahrtrichtung geändert',
                }

                modified_data.append({
                    'Objektklasse': new_data.get('cls', ''),
                    'Kennung': new_data.get('anchor_text', '') or new_data.get('coord_text', ''),
                    'Änderungstyp': field_names.get(field, f'{field} geändert'),
                    'Alt-Wert': old_str,
                    'Neu-Wert': new_str,
                    'Koordinate': new_data.get('coord_text', ''),
                    'Koordinatenwert': _format_coord_value(new_data.get('coord_value')),
                    'Seite': new_data.get('page', ''),
                })

        if modified_data:
            dataframes['Modifiziert'] = pd.DataFrame(modified_data)

    return dataframes


def _format_coord_value(val) -> str:
    """Format coordinate value for display"""
    if val is None:
        return ''
    try:
        if isinstance(val, float):
            import math
            if math.isnan(val):
                return ''
            return f"{val:.4f}"
        return str(val)
    except:
        return str(val) if val else ''
